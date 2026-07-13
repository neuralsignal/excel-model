"""Builders for auxiliary sheets: Assumptions, Drivers, and Inputs."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from excel_model.injection_guard import sanitize_cell_text
from excel_model.loader import InputData
from excel_model.models._sheet_builder import write_four_col_header, write_section_header, write_title_row
from excel_model.named_ranges import register_named_range
from excel_model.spec import AssumptionDef, ModelSpec, ScenarioDef
from excel_model.style import (
    StyleConfig,
    apply_assumption_sheet_validation,
    apply_header_style,
    apply_normal_style,
    get_number_format,
)
from excel_model.time_engine import Period


@dataclass(frozen=True)
class SheetWriteContext:
    """Fixed context for writing rows within a single auxiliary sheet."""

    wb: Workbook
    ws: Worksheet
    sheet_name: str
    style: StyleConfig


def write_assumption_row(
    write_ctx: SheetWriteContext,
    row: int,
    label: str,
    range_name: str,
    value: Any,
    fmt: str,
) -> None:
    """Write a single assumption row: label, named range, value, format."""
    ws = write_ctx.ws
    ws.cell(row=row, column=1, value=sanitize_cell_text(label))
    ws.cell(row=row, column=2, value=range_name)

    value_cell = ws.cell(row=row, column=3, value=value)
    value_cell.number_format = get_number_format(fmt, write_ctx.style)
    value_cell.alignment = Alignment(horizontal="right")

    apply_normal_style(ws.cell(row=row, column=1), write_ctx.style)
    apply_normal_style(ws.cell(row=row, column=2), write_ctx.style)
    apply_normal_style(value_cell, write_ctx.style)

    ws.cell(row=row, column=4, value=fmt)
    apply_normal_style(ws.cell(row=row, column=4), write_ctx.style)

    register_named_range(write_ctx.wb, range_name, write_ctx.sheet_name, row, 3)


def build_assumptions_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    scenario_prefix: str,
) -> dict[str, int]:
    """Build the Assumptions sheet.

    Returns a dict of assumption_name → row number (1-based).
    Registers named ranges for all assumptions.
    """
    sheet_name = "Assumptions"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    write_title_row(ws, f"{spec.title} — Assumptions", 4, style)
    write_four_col_header(ws, style)

    current_row = 3
    assumption_rows: dict[str, int] = {}

    write_ctx = SheetWriteContext(wb=wb, ws=ws, sheet_name=sheet_name, style=style)

    groups: dict[str, list[AssumptionDef]] = {}
    for assumption in spec.assumptions:
        groups.setdefault(assumption.group, []).append(assumption)

    for group_name, assumptions in groups.items():
        write_section_header(ws, group_name, current_row, 4, style)
        current_row += 1

        group_start_row = current_row
        for assumption in assumptions:
            range_name = f"{scenario_prefix}{assumption.name}" if scenario_prefix else assumption.name
            write_assumption_row(
                write_ctx,
                current_row,
                assumption.label,
                range_name,
                assumption.value,
                assumption.format,
            )
            assumption_rows[assumption.name] = current_row
            current_row += 1

        apply_assumption_sheet_validation(
            ws=ws,
            assumptions=assumptions,
            value_col=3,
            format_col=4,
            start_row=group_start_row,
        )

    return assumption_rows


def build_drivers_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    scenarios: tuple[ScenarioDef, ...],
) -> dict[str, int]:
    """Build the Drivers sheet with one section per scenario.

    Returns a dict of driver_name → row number (1-based) for the last section written.
    Registers scenario-prefixed named ranges for each driver.
    """
    sheet_name = "Drivers"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    write_title_row(ws, f"{spec.title} — Drivers", 4, style)
    write_four_col_header(ws, style)

    current_row = 3
    driver_rows: dict[str, int] = {}

    for scenario in scenarios:
        prefix = scenario.name.capitalize()

        write_section_header(ws, f"{scenario.label} Drivers", current_row, 4, style)
        current_row += 1

        for driver in spec.drivers:
            range_name = f"{prefix}{driver.name}"
            # Override value if specified in driver_overrides
            value = scenario.driver_overrides.get(driver.name, driver.value)

            write_assumption_row(
                wb,
                ws,
                sheet_name,
                current_row,
                driver.label,
                range_name,
                value,
                driver.format,
                style,
            )

            driver_rows[driver.name] = current_row
            current_row += 1

    return driver_rows


def build_inputs_sheet(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
    periods: list[Period],
) -> dict[str, int]:
    """Build the Inputs sheet.

    Returns a dict of line_item_key → row number (1-based).
    """
    sheet_name = "Inputs"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    history_periods = [p for p in periods if p.is_history]
    n_cols = 1 + len(history_periods)  # label col + one per history period

    write_title_row(ws, "Historical Input Data", n_cols, style)

    if not history_periods:
        ws["A2"] = "No history periods defined."
        return {}

    header_cell = ws.cell(row=2, column=1, value="Line Item")
    apply_header_style(header_cell, style)
    for col_idx, period in enumerate(history_periods, start=2):
        cell = ws.cell(row=2, column=col_idx, value=sanitize_cell_text(period.label))
        apply_header_style(cell, style)

    ws.column_dimensions["A"].width = 28

    if inputs is None:
        ws["A3"] = "(No input data provided)"
        return {}

    inputs_row_map: dict[str, int] = {}
    current_row = 3

    for line_item_key, source_col in spec.inputs.value_cols.items():
        label_cell = ws.cell(row=current_row, column=1, value=line_item_key)
        apply_normal_style(label_cell, style)

        inputs_row_map[line_item_key] = current_row

        for col_idx, period in enumerate(history_periods, start=2):
            period_val = None
            if inputs.period_col in inputs.df.columns and source_col in inputs.df.columns:
                rows_for_period = inputs.df.filter(inputs.df[inputs.period_col] == period.label)
                if len(rows_for_period) > 0:
                    period_val = rows_for_period[source_col][0]
            cell = ws.cell(row=current_row, column=col_idx, value=period_val)
            apply_normal_style(cell, style)

        current_row += 1

    return inputs_row_map
