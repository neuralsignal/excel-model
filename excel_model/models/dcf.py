"""DCF sheet builder — has its own model sheet builder for NPV_SUM aggregation."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from excel_model.exceptions import ExcelModelError
from excel_model.formula_engine import CellContext, render_formula
from excel_model.loader import InputData
from excel_model.models._sheet_builder import (
    apply_data_cell_style,
    apply_label_style,
    assign_row_map,
    build_assumptions_sheet,
    build_inputs_sheet,
    build_model_header,
    compute_proj_col_range,
    group_line_items_by_section,
    write_history_border,
    write_section_header,
)
from excel_model.named_ranges import get_col_letter
from excel_model.spec import LineItemDef, ModelSpec
from excel_model.style import (
    StyleConfig,
    apply_header_style,
    apply_history_col_style,
    get_number_format,
)
from excel_model.time_engine import Period


def build_dcf(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
    periods: list[Period],
) -> None:
    """Build Assumptions, Inputs, and DCF Model sheets."""
    build_assumptions_sheet(wb, spec, style, "")
    inputs_row_map = build_inputs_sheet(wb, spec, inputs, style, periods)
    _build_dcf_model_sheet(wb, spec, style, periods, inputs_row_map)


def _write_npv_sum_cell(
    ws: Worksheet,
    li: LineItemDef,
    row: int,
    spec: ModelSpec,
    row_map: dict[str, int],
    inputs_row_map: dict[str, int],
    first_proj_col_letter: str,
    last_proj_col_letter: str,
    style: StyleConfig,
) -> None:
    """Write NPV_SUM formula in the first data column, spanning all projection cols."""
    first_data_col = 2
    col_letter = get_col_letter(first_data_col)
    ctx = CellContext(
        period_index=0,
        n_history=spec.n_history_periods,
        row=row,
        col=first_data_col,
        col_letter=col_letter,
        prior_col_letter="",
        named_ranges={a.name: a.name for a in spec.assumptions},
        row_map=row_map,
        inputs_row_map=inputs_row_map,
        scenario_prefix="",
        first_proj_col_letter=first_proj_col_letter,
        last_proj_col_letter=last_proj_col_letter,
        entity_col_range="",
        driver_names=frozenset(),
    )
    value = render_formula(li.formula_type, dict(li.formula_params), ctx)
    cell = ws.cell(row=row, column=first_data_col, value=value)
    cell.number_format = get_number_format("currency", style)
    cell.alignment = Alignment(horizontal="right")
    apply_data_cell_style(cell, li, style, False)


def _write_standard_cells(
    ws: Worksheet,
    li: LineItemDef,
    periods: list[Period],
    spec: ModelSpec,
    row: int,
    row_map: dict[str, int],
    inputs_row_map: dict[str, int],
    first_proj_col_letter: str,
    last_proj_col_letter: str,
    style: StyleConfig,
) -> None:
    """Write standard per-column formula cells for a line item."""
    for col_idx, period in enumerate(periods, start=2):
        col_letter = get_col_letter(col_idx)
        prior_col_letter = get_col_letter(col_idx - 1) if col_idx > 2 else ""

        params = dict(li.formula_params)
        if li.formula_type == "input_ref":
            params["line_item_key"] = li.key

        ctx = CellContext(
            period_index=period.index,
            n_history=spec.n_history_periods,
            row=row,
            col=col_idx,
            col_letter=col_letter,
            prior_col_letter=prior_col_letter,
            named_ranges={a.name: a.name for a in spec.assumptions},
            row_map=row_map,
            inputs_row_map=inputs_row_map,
            scenario_prefix="",
            first_proj_col_letter=first_proj_col_letter,
            last_proj_col_letter=last_proj_col_letter,
            entity_col_range="",
            driver_names=frozenset(),
        )

        value = render_formula(li.formula_type, params, ctx)
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.number_format = get_number_format("currency", style)
        cell.alignment = Alignment(horizontal="right")
        apply_data_cell_style(cell, li, style, period.is_history)


def _build_dcf_model_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    periods: list[Period],
    inputs_row_map: dict[str, int],
) -> None:
    """DCF model sheet — renders NPV_SUM only in the first data column."""
    ws = wb.create_sheet(title="Model")
    total_cols = 1 + len(periods)
    first_proj_col_letter, last_proj_col_letter = compute_proj_col_range(periods, 1, 2)

    build_model_header(ws, spec.title, total_cols, style, "Line Item", 14, "B3")

    # Row 2: Period labels
    for col_idx, period in enumerate(periods, start=2):
        cell = ws.cell(row=2, column=col_idx, value=period.label)
        if period.is_history:
            apply_history_col_style(cell, style)
            cell.font = Font(name=style.font_name, size=style.font_size, bold=True)
        else:
            apply_header_style(cell, style)


    sections_order, sections_items = group_line_items_by_section(spec.line_items)
    row_map = assign_row_map(sections_order, sections_items, 3)

    # Write data
    current_row = 3
    for section in sections_order:
        if section:
            write_section_header(ws, section, current_row, total_cols, style)
            current_row += 1

        for li in sections_items[section]:
            if row_map[li.key] != current_row:
                raise ExcelModelError(f"Row mismatch for {li.key!r}: expected {current_row}, got {row_map[li.key]}")

            label_cell = ws.cell(row=current_row, column=1, value=li.label)
            apply_label_style(label_cell, li, style)

            if li.formula_type == "npv_sum":
                _write_npv_sum_cell(
                    ws,
                    li,
                    current_row,
                    spec,
                    row_map,
                    inputs_row_map,
                    first_proj_col_letter,
                    last_proj_col_letter,
                    style,
                )
            else:
                _write_standard_cells(
                    ws,
                    li,
                    periods,
                    spec,
                    current_row,
                    row_map,
                    inputs_row_map,
                    first_proj_col_letter,
                    last_proj_col_letter,
                    style,
                )

            write_history_border(ws, current_row, spec.n_history_periods, total_cols)
            current_row += 1
