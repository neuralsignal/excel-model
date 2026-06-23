"""Scenario Analysis sheet builder — side-by-side column groups."""

from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_model.formula_engine import render_formula
from excel_model.formula_types import CellContext
from excel_model.injection_guard import sanitize_cell_text
from excel_model.loader import InputData
from excel_model.models._auxiliary_sheets import (
    build_assumptions_sheet,
    build_drivers_sheet,
    build_inputs_sheet,
    write_assumption_row,
)
from excel_model.models._sheet_builder import (
    HeaderLayout,
    apply_label_style,
    assign_row_map,
    build_model_header,
    compute_proj_col_range,
    effective_format,
    group_line_items_by_section,
    maybe_apply_variance_formatting,
    resolve_formula_params,
    write_four_col_header,
    write_grouped_period_headers,
    write_section_header,
    write_title_row,
)
from excel_model.named_ranges import get_col_letter
from excel_model.spec import ModelSpec, ScenarioDef
from excel_model.style import (
    StyleConfig,
    get_number_format,
)
from excel_model.time_engine import Period


def build_scenario(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
    periods: list[Period],
) -> None:
    """Build Assumptions, (optionally Drivers), Inputs, and Scenario Model sheets."""
    if spec.drivers:
        # New mode: separate Assumptions (bare names) and Drivers (prefixed per scenario)
        build_assumptions_sheet(wb, spec, style, "")
        build_drivers_sheet(wb, spec, style, spec.scenarios)
    else:
        # Legacy mode: all assumptions prefixed per scenario
        _build_scenario_assumptions(wb, spec, style)

    inputs_row_map = build_inputs_sheet(wb, spec, inputs, style, periods)
    _build_scenario_model_sheet(wb, spec, style, periods, inputs_row_map)


def _scenario_prefix(scenario: ScenarioDef) -> str:
    """Convert scenario name to CamelCase prefix for named ranges."""
    return scenario.name.capitalize()


def _build_scenario_assumptions(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
) -> None:
    """Build Assumptions sheet with one section per scenario."""
    sheet_name = "Assumptions"
    ws = wb.create_sheet(title=sheet_name)

    write_title_row(ws, f"{spec.title} — Scenario Assumptions", 4, style)
    write_four_col_header(ws, style)

    current_row = 3

    for scenario in spec.scenarios:
        prefix = _scenario_prefix(scenario)

        write_section_header(ws, f"{scenario.label} Assumptions", current_row, 4, style)
        current_row += 1

        for assumption in spec.assumptions:
            range_name = f"{prefix}{assumption.name}"
            value = scenario.assumption_overrides.get(assumption.name, assumption.value)

            write_assumption_row(
                wb, ws, sheet_name, current_row, assumption.label, range_name, value, assumption.format, style
            )
            current_row += 1


def _build_scenario_model_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    periods: list[Period],
    inputs_row_map: dict[str, int],
) -> None:
    """Scenario model: Base | Bull | Bear column groups per period."""
    ws = wb.create_sheet(title="Model")

    n_scenarios = len(spec.scenarios)
    n_sub_cols = n_scenarios
    total_cols = 1 + len(periods) * n_sub_cols

    first_proj_col_letter, last_proj_col_letter = compute_proj_col_range(periods, n_sub_cols, 2)

    build_model_header(ws, spec.title, total_cols, style, HeaderLayout("Line Item", 13, "B4"))
    sub_labels = tuple(sanitize_cell_text(s.label) for s in spec.scenarios)
    write_grouped_period_headers(ws, periods, sub_labels, n_sub_cols, style)

    sections_order, sections_items = group_line_items_by_section(spec.line_items)
    row_map = assign_row_map(sections_order, sections_items, 4)

    current_row = 4
    for section in sections_order:
        if section:
            write_section_header(ws, section, current_row, total_cols, style)
            current_row += 1

        for li in sections_items[section]:
            label_cell = ws.cell(row=current_row, column=1, value=sanitize_cell_text(li.label))
            apply_label_style(label_cell, li, style)

            for p_idx, period in enumerate(periods):
                base_col = 2 + p_idx * n_sub_cols

                for s_idx, scenario in enumerate(spec.scenarios):
                    col_idx = base_col + s_idx
                    col_letter = get_col_letter(col_idx)
                    prior_col_letter = get_col_letter(col_idx - n_sub_cols) if p_idx > 0 else ""
                    prefix = _scenario_prefix(scenario)

                    params = resolve_formula_params(li)

                    all_named = {a.name: a.name for a in spec.assumptions}
                    for d in spec.drivers:
                        all_named[d.name] = d.name

                    ctx = CellContext(
                        period_index=period.index,
                        n_history=spec.n_history_periods,
                        row=current_row,
                        col=col_idx,
                        col_letter=col_letter,
                        prior_col_letter=prior_col_letter,
                        named_ranges=all_named,
                        row_map=row_map,
                        inputs_row_map=inputs_row_map,
                        scenario_prefix=prefix,
                        first_proj_col_letter=first_proj_col_letter,
                        last_proj_col_letter=last_proj_col_letter,
                        entity_col_range="",
                        driver_names=frozenset(d.name for d in spec.drivers),
                    )

                    value = render_formula(li.formula_type, params, ctx)
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.number_format = get_number_format(effective_format(li), style)
                    cell.alignment = Alignment(horizontal="right")
                    apply_label_style(cell, li, style)

            maybe_apply_variance_formatting(ws, li, current_row, total_cols, style)

            current_row += 1
