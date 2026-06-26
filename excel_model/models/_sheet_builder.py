"""Shared utilities for building model sheet layouts."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.cell import Cell
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_model.formula_types import CellContext
from excel_model.injection_guard import sanitize_cell_text
from excel_model.named_ranges import get_col_letter
from excel_model.spec import LineItemDef
from excel_model.style import (
    StyleConfig,
    apply_conditional_formatting,
    apply_header_style,
    apply_history_col_style,
    apply_normal_style,
    apply_section_header_style,
    apply_subtotal_style,
    apply_total_style,
)
from excel_model.time_engine import Period


@dataclass(frozen=True)
class SheetRenderContext:
    """Shared rendering state passed to model sheet cell-writing functions."""

    row_map: dict[str, int]
    inputs_row_map: dict[str, int]
    first_proj_col_letter: str
    last_proj_col_letter: str
    n_history: int
    named_ranges: dict[str, str]
    style: StyleConfig


@dataclass(frozen=True)
class HeaderLayout:
    """Layout parameters for the standard model sheet header."""

    label_col_header: str
    data_col_width: int
    freeze_cell: str


def build_model_header(
    ws: Worksheet,
    title: str,
    total_cols: int,
    style: StyleConfig,
    layout: HeaderLayout,
) -> None:
    """Build the standard model sheet header rows shared by all builders.

    Handles: row-1 title merge + style, row-2 label column header,
    column-A width, data-column widths, and freeze panes.
    """
    write_title_row(ws, title, total_cols, style)

    label_header = ws.cell(row=2, column=1, value=layout.label_col_header)
    apply_header_style(label_header, style)

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = layout.data_col_width

    ws.freeze_panes = layout.freeze_cell


def group_line_items_by_section(
    line_items: tuple[LineItemDef, ...],
) -> tuple[list[str], dict[str, list[LineItemDef]]]:
    """Group line items by section, preserving insertion order.

    Returns (sections_order, sections_items) where sections_order is the list
    of unique section names in first-seen order and sections_items maps each
    section name to its line items.
    """
    sections_order: list[str] = []
    sections_items: dict[str, list[LineItemDef]] = {}
    for li in line_items:
        if li.section not in sections_items:
            sections_order.append(li.section)
            sections_items[li.section] = []
        sections_items[li.section].append(li)
    return sections_order, sections_items


_FOUR_COL_HEADERS = ("Parameter", "Named Range", "Value", "Format")
_FOUR_COL_WIDTHS = (30, 25, 15, 12)


def write_four_col_header(ws: Worksheet, style: StyleConfig) -> None:
    """Write the standard 4-column header row and set column widths.

    Used by Assumptions, Drivers, and Scenario Assumptions sheets.
    """
    col_letters = ("A", "B", "C", "D")
    for col_idx, header in enumerate(_FOUR_COL_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell, style)
    for letter, width in zip(col_letters, _FOUR_COL_WIDTHS, strict=True):
        ws.column_dimensions[letter].width = width


def write_title_row(ws: Worksheet, title: str, total_cols: int, style: StyleConfig) -> None:
    """Write merged title row (row 1) with header styling."""
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    cell = ws["A1"]
    cell.value = sanitize_cell_text(title)
    apply_header_style(cell, style)
    ws.row_dimensions[1].height = 20


def assign_row_map(
    sections_order: list[str],
    sections_items: dict[str, list[LineItemDef]],
    start_row: int,
) -> dict[str, int]:
    """First-pass row number assignment for all line items."""
    current_row = start_row
    row_map: dict[str, int] = {}
    for section in sections_order:
        if section:
            current_row += 1
        for li in sections_items[section]:
            row_map[li.key] = current_row
            current_row += 1
    return row_map


def write_section_header(
    ws: Worksheet,
    section: str,
    row: int,
    total_cols: int,
    style: StyleConfig,
) -> None:
    """Write a merged section header row."""
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws[f"A{row}"].value = sanitize_cell_text(section)
    apply_section_header_style(ws[f"A{row}"], style)


def apply_label_style(cell: Cell, li: LineItemDef, style: StyleConfig) -> None:
    """Apply normal/subtotal/total style to a line item label cell."""
    apply_normal_style(cell, style)
    if li.is_subtotal:
        apply_subtotal_style(cell, style)
    elif li.is_total:
        apply_total_style(cell, style)


def apply_data_cell_style(cell: Cell, li: LineItemDef, style: StyleConfig, is_history: bool) -> None:
    """Apply appropriate style to a data cell based on line item type and history status."""
    if li.is_subtotal:
        apply_subtotal_style(cell, style)
    elif li.is_total:
        apply_total_style(cell, style)
    else:
        apply_normal_style(cell, style)
        if is_history:
            apply_history_col_style(cell, style)


def compute_proj_col_range(
    periods: list[Period],
    col_multiplier: int,
    col_offset: int,
) -> tuple[str, str]:
    """Compute first/last projection column letters for formula context."""
    proj = [p for p in periods if not p.is_history]
    if not proj:
        return "", ""
    return (
        get_col_letter(col_offset + proj[0].index * col_multiplier),
        get_col_letter(col_offset + proj[-1].index * col_multiplier + col_multiplier - 1),
    )


def write_grouped_period_headers(
    ws: Worksheet,
    periods: list[Period],
    sub_labels: tuple[str, ...],
    n_sub_cols: int,
    style: StyleConfig,
) -> None:
    """Write merged period group headers (row 2) and per-sub-column labels (row 3)."""
    label_header = ws.cell(row=2, column=1, value="Line Item")
    apply_header_style(label_header, style)
    for p_idx, period in enumerate(periods):
        base_col = 2 + p_idx * n_sub_cols
        end_col = base_col + n_sub_cols - 1
        ws.merge_cells(f"{get_column_letter(base_col)}2:{get_column_letter(end_col)}2")
        ph = ws.cell(row=2, column=base_col, value=sanitize_cell_text(period.label))
        apply_header_style(ph, style)

    sub_label_cell = ws.cell(row=3, column=1, value="")
    apply_header_style(sub_label_cell, style)
    for p_idx in range(len(periods)):
        base_col = 2 + p_idx * n_sub_cols
        for s_idx, label in enumerate(sub_labels):
            cell = ws.cell(row=3, column=base_col + s_idx, value=label)
            apply_header_style(cell, style)


def effective_format(li: LineItemDef) -> str:
    """Return li.format, defaulting to 'currency'."""
    return li.format if li.format else "currency"


def resolve_formula_params(li: LineItemDef) -> dict[str, Any]:
    """Return a copy of formula_params with input_ref key injected when needed."""
    params = dict(li.formula_params)
    if li.formula_type == "input_ref":
        params["line_item_key"] = li.key
    return params


def maybe_apply_variance_formatting(
    ws: Worksheet,
    li: LineItemDef,
    current_row: int,
    total_cols: int,
    style: StyleConfig,
) -> None:
    """Apply conditional formatting to variance rows if positive_is_good is set."""
    if li.formula_type in ("variance", "variance_pct") and "positive_is_good" in li.formula_params:
        positive_is_good = bool(li.formula_params["positive_is_good"])
        cf_range = f"{get_column_letter(2)}{current_row}:{get_column_letter(total_cols)}{current_row}"
        apply_conditional_formatting(ws, cf_range, positive_is_good, style)


def write_history_border(ws: Worksheet, row: int, n_history: int, total_cols: int) -> None:
    """Write thin vertical border at the first projection column."""
    if n_history > 0:
        border_col = 2 + n_history
        if border_col <= total_cols:
            ws.cell(row=row, column=border_col).border = Border(left=Side(style="thin"))


@dataclass(frozen=True)
class CellPosition:
    """Per-cell positional data for formula rendering."""

    period_index: int
    col: int
    col_letter: str
    prior_col_letter: str
    row: int


def make_cell_context(
    render_ctx: SheetRenderContext,
    cell_pos: CellPosition,
    scenario_prefix: str,
    entity_col_range: str,
    driver_names: frozenset[str],
) -> CellContext:
    """Build a CellContext from shared render state plus per-cell parameters."""
    return CellContext(
        period_index=cell_pos.period_index,
        n_history=render_ctx.n_history,
        row=cell_pos.row,
        col=cell_pos.col,
        col_letter=cell_pos.col_letter,
        prior_col_letter=cell_pos.prior_col_letter,
        named_ranges=render_ctx.named_ranges,
        row_map=render_ctx.row_map,
        inputs_row_map=render_ctx.inputs_row_map,
        scenario_prefix=scenario_prefix,
        first_proj_col_letter=render_ctx.first_proj_col_letter,
        last_proj_col_letter=render_ctx.last_proj_col_letter,
        entity_col_range=entity_col_range,
        driver_names=driver_names,
    )
