"""Comparison model builder — entities as columns (not time periods)."""

from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_model.exceptions import ExcelModelError
from excel_model.formula_engine import CellContext, render_formula
from excel_model.loader import InputData
from excel_model.models._auxiliary_sheets import build_assumptions_sheet
from excel_model.models._sheet_builder import (
    apply_data_cell_style,
    apply_label_style,
    assign_row_map,
    build_model_header,
    group_line_items_by_section,
    write_section_header,
)
from excel_model.named_ranges import get_col_letter
from excel_model.spec import ModelSpec
from excel_model.style import (
    StyleConfig,
    apply_header_style,
    get_number_format,
)


def build_comparison(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
) -> None:
    """Build Assumptions and Comparison Model sheets."""
    build_assumptions_sheet(wb, spec, style, "")
    _build_comparison_model_sheet(wb, spec, style)


def _write_entity_headers(
    ws: object,
    spec: ModelSpec,
    style: StyleConfig,
) -> None:
    """Write entity label headers (row 2)."""
    label_header = ws.cell(row=2, column=1, value="Metric")  # type: ignore[union-attr]
    apply_header_style(label_header, style)
    for e_idx, entity in enumerate(spec.entities):
        cell = ws.cell(row=2, column=2 + e_idx, value=entity.label)  # type: ignore[union-attr]
        apply_header_style(cell, style)


def _build_comparison_model_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
) -> None:
    """Comparison model: entities as columns, line items as rows."""
    ws = wb.create_sheet(title="Model")

    entities = spec.entities
    n_entities = len(entities)
    total_cols = 1 + n_entities

    build_model_header(ws, spec.title, total_cols, style, "Metric", 16, "B3")
    _write_entity_headers(ws, spec, style)

    # Build entity_col_range for RANK/MAX formulas
    first_entity_col = get_col_letter(2)
    last_entity_col = get_col_letter(1 + n_entities)

    sections_order, sections_items = group_line_items_by_section(spec.line_items)
    row_map = assign_row_map(sections_order, sections_items, 3)

    # Write data
    current_row = 3
    for section in sections_order:
        if section:
            write_section_header(ws, section, current_row, total_cols, style)
            current_row += 1

        for li in sections_items[section]:
            if row_map[li.key] != current_row:
                raise ExcelModelError(f"Row mismatch for {li.key!r}: expected {current_row}, got {row_map[li.key]}")

            label_cell = ws.cell(row=current_row, column=1, value=li.label)
            apply_label_style(label_cell, li, style)

            entity_col_range = f"${first_entity_col}${current_row}:${last_entity_col}${current_row}"

            for e_idx, _entity in enumerate(entities):
                col_idx = 2 + e_idx
                col_letter = get_col_letter(col_idx)

                params = dict(li.formula_params)

                # For index_to_base, inject the base entity's column letter
                if li.formula_type == "index_to_base":
                    base_entity_key = params.get("base_entity_key", "")
                    base_col_idx = _entity_col_index(spec, base_entity_key)
                    if base_col_idx is not None:
                        params["_base_col_letter"] = get_col_letter(base_col_idx)

                ctx = CellContext(
                    period_index=0,
                    n_history=0,
                    row=current_row,
                    col=col_idx,
                    col_letter=col_letter,
                    prior_col_letter="",
                    named_ranges={a.name: a.name for a in spec.assumptions},
                    row_map=row_map,
                    inputs_row_map={},
                    scenario_prefix="",
                    first_proj_col_letter="",
                    last_proj_col_letter="",
                    entity_col_range=entity_col_range,
                    driver_names=frozenset(),
                )

                value = render_formula(li.formula_type, params, ctx)
                cell = ws.cell(row=current_row, column=col_idx, value=value)

                fmt = "percent" if li.formula_type in ("ratio", "index_to_base") else "currency"
                cell.number_format = get_number_format(fmt, style)
                cell.alignment = Alignment(horizontal="right")
                apply_data_cell_style(cell, li, style, False)

            current_row += 1


def _entity_col_index(spec: ModelSpec, entity_key: str) -> int | None:
    """Return the 1-based column index for the given entity key, or None."""
    for i, entity in enumerate(spec.entities):
        if entity.key == entity_key:
            return 2 + i
    return None
