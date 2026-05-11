"""Builder functions for tabular data sheets and SUMIFS pivot sheets.

Follows the spec-driven pattern: callers construct a DataSheetDef or
SumifsPivotDef, which is validated then built into an openpyxl Worksheet.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_model.data_sheet_validator import (
    validate_data_sheet_def,
    validate_sumifs_pivot_def,
)
from excel_model.exceptions import SpecValidationError
from excel_model.models._sheet_builder import write_title_row
from excel_model.spec import DataSheetDef, SumifsPivotDef
from excel_model.style import (
    StyleConfig,
    apply_alt_row_style,
    apply_header_style,
    apply_normal_style,
)


@dataclass(frozen=True)
class RowWriteContext:
    """Shared context for writing cells in a single data row."""

    ws: Worksheet
    row_idx: int
    is_alt: bool
    style: StyleConfig


def _quote_sheet_ref(name: str) -> str:
    """Return name single-quoted for safe use in an Excel formula sheet reference.

    Excel requires single quotes around sheet names that contain spaces or other
    non-alphanumeric characters (e.g. ``'My Sheet'!A1``). The validator already
    rejects single quotes and other unsafe characters, so unconditional quoting
    is safe and keeps formulas consistent for every sheet name.
    """
    return f"'{name}'"


def build_data_sheet(
    wb: Workbook,
    spec: DataSheetDef,
    rows: list[list[Any]],
    style: StyleConfig,
) -> Worksheet:
    """Build a formatted tabular data sheet into wb and return the worksheet.

    Validates the spec before building. Raises SpecValidationError on invalid input.

    Sheet layout:
      Row 1 — merged title (via write_title_row)
      Row 2 — column headers (bold, header fill)
      Row 3+ — data rows (alternating fill on even rows)

    Args:
        wb: target workbook (sheet is appended)
        spec: DataSheetDef with sheet configuration
        rows: list of data rows; each row is a list aligned with spec.headers
        style: StyleConfig from load_style()

    Returns:
        The created Worksheet.
    """
    errors = validate_data_sheet_def(spec)
    if errors:
        raise SpecValidationError(f"Invalid DataSheetDef: {'; '.join(errors)}")

    n_cols = len(spec.headers)
    ws = wb.create_sheet(title=spec.sheet_name)

    write_title_row(ws, spec.title, n_cols, style)

    for col_idx, header in enumerate(spec.headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell, style)

    for row_idx, row_data in enumerate(rows, start=3):
        is_alt = row_idx % 2 == 0
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_normal_style(cell, style)
            if is_alt:
                apply_alt_row_style(cell, style)
            zero_col = col_idx - 1
            if zero_col in spec.number_formats:
                cell.number_format = spec.number_formats[zero_col]

    for col_idx, width in enumerate(spec.col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{spec.freeze_row + 1}"

    return ws


def build_sumifs_pivot(
    wb: Workbook,
    spec: SumifsPivotDef,
    row_labels: list[list[Any]],
    style: StyleConfig,
) -> Worksheet:
    """Build a pivot-style sheet where each data cell is a SUMIFS formula.

    Validates the spec before building. Raises SpecValidationError on invalid input.

    Sheet layout:
      Row 1 — merged title
      Row 2 — row_label_headers + col_dim_values + optional Total + optional YoY headers
      Row 3+ — one row per entry in row_labels

    Args:
        wb: target workbook
        spec: SumifsPivotDef with sheet configuration
        row_labels: one list per data row; each inner list has len(spec.row_label_headers) values
        style: StyleConfig from load_style()

    Returns:
        The created Worksheet.
    """
    errors = validate_sumifs_pivot_def(spec)
    if errors:
        raise SpecValidationError(f"Invalid SumifsPivotDef: {'; '.join(errors)}")

    n_label_cols = len(spec.row_label_headers)
    n_data_cols = len(spec.col_dim_values)

    header_row: list[Any] = list(spec.row_label_headers) + list(spec.col_dim_values)
    if spec.append_total:
        header_row.append("Total")
    if spec.append_yoy:
        for i in range(n_data_cols - 1):
            header_row.append(f"YoY {spec.col_dim_values[i]}\u2192{spec.col_dim_values[i + 1]}")

    n_total_cols = len(header_row)
    ws = wb.create_sheet(title=spec.sheet_name)

    write_title_row(ws, spec.title, n_total_cols, style)

    for col_idx, hdr in enumerate(header_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        apply_header_style(cell, style)

    first_data_col = n_label_cols + 1
    last_data_col = n_label_cols + n_data_cols

    for row_idx, label_vals in enumerate(row_labels, start=3):
        ctx = RowWriteContext(ws, row_idx, row_idx % 2 == 0, style)

        for label_col_idx, label_val in enumerate(label_vals, start=1):
            cell = ws.cell(row=row_idx, column=label_col_idx, value=label_val)
            apply_normal_style(cell, style)
            if ctx.is_alt:
                apply_alt_row_style(cell, style)

        _write_sumifs_data_cells(ctx, spec, first_data_col, n_data_cols)

        if spec.append_total:
            _write_total_cell(
                ctx,
                last_data_col + 1,
                first_data_col,
                last_data_col,
                spec.number_format_data,
            )

        if spec.append_yoy:
            yoy_start_col = last_data_col + (2 if spec.append_total else 1)
            _write_yoy_cells(ctx, yoy_start_col, first_data_col, n_data_cols, spec.number_format_pct)

    for col_idx, width in enumerate(spec.col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{spec.freeze_row + 1}"

    return ws


def _write_sumifs_data_cells(
    ctx: RowWriteContext,
    spec: SumifsPivotDef,
    first_data_col: int,
    n_data_cols: int,
) -> None:
    """Write SUMIFS formula cells for one data row."""
    sheet_ref = _quote_sheet_ref(spec.data_sheet)
    for data_col_offset in range(n_data_cols):
        data_col_idx = first_data_col + data_col_offset
        data_col_letter = get_column_letter(data_col_idx)

        row_criteria_parts: list[str] = []
        for filter_idx, filter_data_col in enumerate(spec.row_filter_cols):
            label_col_letter = get_column_letter(filter_idx + 1)
            row_criteria_parts.append(
                f"{sheet_ref}!${filter_data_col}:${filter_data_col},${label_col_letter}{ctx.row_idx}"
            )

        col_dim_criterion = f"{sheet_ref}!${spec.col_filter_col}:${spec.col_filter_col},{data_col_letter}$2"

        all_criteria = ",".join(row_criteria_parts + [col_dim_criterion])
        formula = f"=SUMIFS({sheet_ref}!${spec.value_col}:${spec.value_col},{all_criteria})"

        cell = ctx.ws.cell(row=ctx.row_idx, column=data_col_idx, value=formula)
        apply_normal_style(cell, ctx.style)
        if ctx.is_alt:
            apply_alt_row_style(cell, ctx.style)
        cell.number_format = spec.number_format_data


def _write_total_cell(
    ctx: RowWriteContext,
    total_col_idx: int,
    first_data_col: int,
    last_data_col: int,
    number_format: str,
) -> None:
    """Write a SUM total cell for one data row."""
    first_letter = get_column_letter(first_data_col)
    last_letter = get_column_letter(last_data_col)
    formula = f"=SUM({first_letter}{ctx.row_idx}:{last_letter}{ctx.row_idx})"
    cell = ctx.ws.cell(row=ctx.row_idx, column=total_col_idx, value=formula)
    apply_normal_style(cell, ctx.style)
    if ctx.is_alt:
        apply_alt_row_style(cell, ctx.style)
    cell.number_format = number_format


def _write_yoy_cells(
    ctx: RowWriteContext,
    yoy_start_col: int,
    first_data_col: int,
    n_data_cols: int,
    number_format: str,
) -> None:
    """Write YoY percentage change cells for one data row."""
    for yoy_idx in range(n_data_cols - 1):
        prev_col_letter = get_column_letter(first_data_col + yoy_idx)
        cur_col_letter = get_column_letter(first_data_col + yoy_idx + 1)
        yoy_col_idx = yoy_start_col + yoy_idx
        yoy_formula = (
            f"=IF({prev_col_letter}{ctx.row_idx}=0,0,"
            f"({cur_col_letter}{ctx.row_idx}-{prev_col_letter}{ctx.row_idx})"
            f"/ABS({prev_col_letter}{ctx.row_idx}))"
        )
        cell = ctx.ws.cell(row=ctx.row_idx, column=yoy_col_idx, value=yoy_formula)
        apply_normal_style(cell, ctx.style)
        if ctx.is_alt:
            apply_alt_row_style(cell, ctx.style)
        cell.number_format = number_format
