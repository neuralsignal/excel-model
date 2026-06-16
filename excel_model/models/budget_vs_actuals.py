"""Budget vs Actuals sheet builder."""

from openpyxl import Workbook
from openpyxl.styles import Alignment

from excel_model.formula_engine import render_formula
from excel_model.formula_types import CellContext
from excel_model.injection_guard import sanitize_cell_text
from excel_model.loader import InputData
from excel_model.models._auxiliary_sheets import build_assumptions_sheet, build_inputs_sheet
from excel_model.models._sheet_builder import (
    HeaderLayout,
    apply_label_style,
    assign_row_map,
    build_model_header,
    compute_proj_col_range,
    effective_format,
    group_line_items_by_section,
    maybe_apply_variance_formatting,
    resolve_formula_params,
    write_grouped_period_headers,
    write_section_header,
)
from excel_model.named_ranges import get_col_letter
from excel_model.spec import ModelSpec
from excel_model.style import (
    StyleConfig,
    get_number_format,
)
from excel_model.time_engine import Period


def build_budget_vs_actuals(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
    periods: list[Period],
) -> None:
    """Build Assumptions, Inputs, and Budget vs Actuals Model sheets."""
    build_assumptions_sheet(wb, spec, style, "")
    inputs_row_map = build_inputs_sheet(wb, spec, inputs, style, periods)
    _build_bva_model_sheet(wb, spec, style, periods, inputs_row_map)


def _build_bva_model_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    periods: list[Period],
    inputs_row_map: dict[str, int],
) -> None:
    """BvA: Plan | Actual | Variance | Variance% columns per period."""
    ws = wb.create_sheet(title="Model")

    groups = spec.column_groups
    n_sub_cols = len(groups)
    total_cols = 1 + len(periods) * n_sub_cols

    first_proj_col_letter, last_proj_col_letter = compute_proj_col_range(periods, n_sub_cols, 2)

    build_model_header(ws, spec.title, total_cols, style, HeaderLayout("Line Item", 12, "B4"))
    sub_labels = tuple(sanitize_cell_text(g.label) for g in groups)
    write_grouped_period_headers(ws, periods, sub_labels, n_sub_cols, style)

    sections_order, sections_items = group_line_items_by_section(spec.line_items)
    row_map = assign_row_map(sections_order, sections_items, 4)

    current_row = 4
    for section in sections_order:
        if section:
            write_section_header(ws, section, current_row, total_cols, style)
            current_row += 1

        for li in sections_items[section]:
            label_cell = ws.cell(row=current_row, column=1, value=sanitize_cell_text(li.label))
            apply_label_style(label_cell, li, style)

            for p_idx, period in enumerate(periods):
                base_col = 2 + p_idx * n_sub_cols

                for g_idx, _group in enumerate(groups):
                    col_idx = base_col + g_idx
                    col_letter = get_col_letter(col_idx)
                    prior_col_letter = get_col_letter(col_idx - n_sub_cols) if p_idx > 0 else ""

                    params = resolve_formula_params(li)

                    ctx = CellContext(
                        period_index=period.index,
                        n_history=spec.n_history_periods,
                        row=current_row,
                        col=col_idx,
                        col_letter=col_letter,
                        prior_col_letter=prior_col_letter,
                        named_ranges={a.name: a.name for a in spec.assumptions},
                        row_map=row_map,
                        inputs_row_map=inputs_row_map,
                        scenario_prefix="",
                        first_proj_col_letter=first_proj_col_letter,
                        last_proj_col_letter=last_proj_col_letter,
                        entity_col_range="",
                        driver_names=frozenset(),
                    )

                    value = render_formula(li.formula_type, params, ctx)
                    cell = ws.cell(row=current_row, column=col_idx, value=value)

                    cell.number_format = get_number_format(effective_format(li), style)
                    cell.alignment = Alignment(horizontal="right")
                    apply_label_style(cell, li, style)

            maybe_apply_variance_formatting(ws, li, current_row, total_cols, style)

            current_row += 1
