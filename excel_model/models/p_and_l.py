"""P&L sheet builder."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from excel_model.exceptions import ExcelModelError
from excel_model.formula_engine import CellContext, render_formula
from excel_model.loader import InputData
from excel_model.models._sheet_builder import (
    build_assumptions_sheet,
    build_inputs_sheet,
    build_model_header,
    group_line_items_by_section,
)
from excel_model.named_ranges import get_col_letter
from excel_model.spec import ModelSpec
from excel_model.style import (
    StyleConfig,
    apply_header_style,
    apply_history_col_style,
    apply_normal_style,
    apply_section_header_style,
    apply_subtotal_style,
    apply_total_style,
    get_number_format,
)
from excel_model.time_engine import Period


def build_p_and_l(
    wb: Workbook,
    spec: ModelSpec,
    inputs: InputData | None,
    style: StyleConfig,
    periods: list[Period],
) -> None:
    """Build Assumptions, Inputs, and P&L Model sheets."""
    # 1. Assumptions sheet
    build_assumptions_sheet(wb, spec, style, "")

    # 2. Inputs sheet
    inputs_row_map = build_inputs_sheet(wb, spec, inputs, style, periods)

    # 3. P&L Model sheet
    _build_model_sheet(wb, spec, style, periods, inputs_row_map)


def _build_model_sheet(
    wb: Workbook,
    spec: ModelSpec,
    style: StyleConfig,
    periods: list[Period],
    inputs_row_map: dict[str, int],
) -> None:
    ws = wb.create_sheet(title="Model")

    n_period_cols = len(periods)
    total_cols = 1 + n_period_cols

    # Compute projection column range
    proj_periods = [p for p in periods if not p.is_history]
    if proj_periods:
        first_proj_col_letter = get_col_letter(2 + proj_periods[0].index)
        last_proj_col_letter = get_col_letter(2 + proj_periods[-1].index)
    else:
        first_proj_col_letter = ""
        last_proj_col_letter = ""

    build_model_header(ws, spec.title, total_cols, style, "Line Item", 14, "B3")

    # Row 2: Period labels
    for col_idx, period in enumerate(periods, start=2):
        cell = ws.cell(row=2, column=col_idx, value=period.label)
        if period.is_history:
            apply_history_col_style(cell, style)
            cell.font = Font(name=style.font_name, size=style.font_size, bold=True)
        else:
            apply_header_style(cell, style)

    # First pass: assign row numbers to all line items
    current_row = 3
    row_map: dict[str, int] = {}

    # Group by section in order
    sections_order, sections_items = group_line_items_by_section(spec.line_items)

    # Assign rows (including section header rows)
    for section in sections_order:
        if section:
            current_row += 1  # section header row
        for li in sections_items[section]:
            row_map[li.key] = current_row
            current_row += 1

    # Second pass: write data
    current_row = 3
    for section in sections_order:
        if section:
            # Section header row
            ws.merge_cells(f"A{current_row}:{get_column_letter(total_cols)}{current_row}")
            sec_cell = ws[f"A{current_row}"]
            sec_cell.value = section
            apply_section_header_style(sec_cell, style)
            current_row += 1

        for li in sections_items[section]:
            if row_map[li.key] != current_row:
                raise ExcelModelError(f"Row mismatch for {li.key!r}: expected {current_row}, got {row_map[li.key]}")

            label_cell = ws.cell(row=current_row, column=1, value=li.label)
            apply_normal_style(label_cell, style)
            if li.is_subtotal:
                apply_subtotal_style(label_cell, style)
            elif li.is_total:
                apply_total_style(label_cell, style)

            for col_idx, period in enumerate(periods, start=2):
                col_letter = get_col_letter(col_idx)
                prior_col_letter = get_col_letter(col_idx - 1) if col_idx > 2 else ""

                # For input_ref, we need to pass the line_item_key via formula_params
                params = dict(li.formula_params)
                if li.formula_type == "input_ref":
                    params["line_item_key"] = li.key

                ctx = CellContext(
                    period_index=period.index,
                    n_history=spec.n_history_periods,
                    row=current_row,
                    col=col_idx,
                    col_letter=col_letter,
                    prior_col_letter=prior_col_letter,
                    named_ranges={a.name: a.name for a in spec.assumptions},
                    row_map=row_map,
                    inputs_row_map=inputs_row_map,
                    scenario_prefix="",
                    first_proj_col_letter=first_proj_col_letter,
                    last_proj_col_letter=last_proj_col_letter,
                    entity_col_range="",
                    driver_names=frozenset(),
                )

                value = render_formula(li.formula_type, params, ctx)

                cell = ws.cell(row=current_row, column=col_idx, value=value)
                # Number format: default to currency for most line items
                cell.number_format = get_number_format("currency", style)
                cell.alignment = Alignment(horizontal="right")

                if period.is_history:
                    apply_history_col_style(cell, style)
                if li.is_subtotal:
                    apply_subtotal_style(cell, style)
                elif li.is_total:
                    apply_total_style(cell, style)
                else:
                    apply_normal_style(cell, style)
                    if period.is_history:
                        apply_history_col_style(cell, style)

            # Thin vertical border after last history col
            if spec.n_history_periods > 0:
                border_col = 1 + spec.n_history_periods + 1  # first projection col
                if border_col <= total_cols:
                    bc = ws.cell(row=current_row, column=border_col)
                    bc.border = Border(left=Side(style="thin"))

            current_row += 1
