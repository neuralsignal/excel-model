"""Standalone sheet builders for tabular and SUMIFS-pivot data sheets.

Reuses StyleConfig and openpyxl styling from style.py.
No model spec required — accepts plain Python lists for headers and rows.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_model.models._sheet_builder import write_title_row
from excel_model.style import StyleConfig, apply_header_style, apply_normal_style


_ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")


def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    style: StyleConfig,
    title: str,
    col_widths: list[float],
    number_formats: dict[int, str],
    freeze_row: int,
) -> Worksheet:
    """Write a formatted tabular data sheet into wb and return the worksheet.

    Sheet layout:
      Row 1 — merged title (via write_title_row)
      Row 2 — column headers (bold, header fill)
      Row 3+ — data rows (alternating light grey on odd data rows)

    Args:
        wb: target workbook (sheet is appended)
        sheet_name: Excel tab name
        headers: column header strings
        rows: list of data rows; each row is a list aligned with headers
        style: StyleConfig (font, colors) from load_style()
        title: text for the merged title row (row 1)
        col_widths: column widths in Excel units; index 0 = column A
        number_formats: mapping of 0-based column index → openpyxl format string
        freeze_row: rows 1..freeze_row are frozen; data starts scrolling at freeze_row+1
                    (pass 2 to freeze title + header rows)

    Returns:
        The created Worksheet.
    """
    n_cols = len(headers)
    ws = wb.create_sheet(title=sheet_name)

    write_title_row(ws, title, n_cols, style)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_header_style(cell, style)

    for row_idx, row_data in enumerate(rows, start=3):
        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_normal_style(cell, style)
            if is_alt:
                cell.fill = _ALT_ROW_FILL
            zero_col = col_idx - 1
            if zero_col in number_formats:
                cell.number_format = number_formats[zero_col]

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{freeze_row + 1}"

    return ws


def write_sumifs_pivot(
    wb: Workbook,
    sheet_name: str,
    title: str,
    style: StyleConfig,
    row_label_headers: list[str],
    row_labels: list[list[Any]],
    col_dim_values: list[Any],
    data_sheet: str,
    value_col: str,
    row_filter_cols: list[str],
    col_filter_col: str,
    append_total: bool,
    append_yoy: bool,
    col_widths: list[float],
    number_format_data: str,
    number_format_pct: str,
    freeze_row: int,
) -> Worksheet:
    """Write a pivot-style sheet where each data cell is a SUMIFS formula.

    Sheet layout:
      Row 1 — merged title
      Row 2 — row_label_headers + col_dim_values + optional "Total" + optional YoY headers
      Row 3+ — one row per entry in row_labels

    The SUMIFS formula for each data cell references the data_sheet by column letter:
      =SUMIFS(data_sheet!$value_col:$value_col,
              data_sheet!$row_filter_cols[0]:$..., $A{row},
              data_sheet!$row_filter_cols[1]:$..., $B{row},  (if 2 filter cols)
              data_sheet!$col_filter_col:$..., {year_col}$2)

    Only the first len(row_filter_cols) label columns participate in SUMIFS criteria.
    Remaining label columns (e.g. a "Category" display column) are written statically.

    Args:
        wb: target workbook
        sheet_name: Excel tab name
        title: merged title row text
        style: StyleConfig from load_style()
        row_label_headers: display headers for the entity label columns (e.g. ["Hub"] or ["Hub","Account"])
        row_labels: one list per data row; each inner list has len(row_label_headers) values
        col_dim_values: values for the column dimension (e.g. [2023, 2024, 2025, 2026])
        data_sheet: name of the source data sheet (e.g. "TRANSACTIONS_LNFW")
        value_col: Excel column letter for the summed value (e.g. "AO")
        row_filter_cols: Excel column letters matched against the first N label columns
                         (len <= len(row_label_headers)); e.g. ["AM"] or ["AM","AN"] or ["AQ"]
        col_filter_col: Excel column letter matched against col_dim_values (e.g. "AJ")
        append_total: if True, add a "Total" column = SUM of data cols
        append_yoy: if True, add YoY % columns after data cols (one between each consecutive pair)
        col_widths: column widths; must cover all columns including label, data, total, yoy
        number_format_data: openpyxl format string for data cells (e.g. "#,##0")
        number_format_pct: openpyxl format string for YoY cells (e.g. "0.0%")
        freeze_row: rows 1..freeze_row are frozen

    Returns:
        The created Worksheet.
    """
    n_label_cols = len(row_label_headers)
    n_data_cols = len(col_dim_values)
    n_filter_cols = len(row_filter_cols)

    # Build header row 2: label headers + year values + optional total + optional yoy
    header_row: list[Any] = list(row_label_headers) + list(col_dim_values)
    if append_total:
        header_row.append("Total")
    if append_yoy:
        for i in range(n_data_cols - 1):
            header_row.append(
                f"YoY {col_dim_values[i]}→{col_dim_values[i + 1]}"
            )

    n_total_cols = len(header_row)
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: merged title
    write_title_row(ws, title, n_total_cols, style)

    # Row 2: headers
    for col_idx, hdr in enumerate(header_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        apply_header_style(cell, style)

    # Column index helpers (1-based)
    first_data_col = n_label_cols + 1           # first year column
    last_data_col = n_label_cols + n_data_cols   # last year column

    # Rows 3+: data rows
    for row_idx, label_vals in enumerate(row_labels, start=3):
        is_alt = (row_idx % 2 == 0)

        # Write label columns (static)
        for label_col_idx, label_val in enumerate(label_vals, start=1):
            cell = ws.cell(row=row_idx, column=label_col_idx, value=label_val)
            apply_normal_style(cell, style)
            if is_alt:
                cell.fill = _ALT_ROW_FILL

        # Write SUMIFS formula for each data column
        for data_col_offset, _year_val in enumerate(col_dim_values):
            data_col_idx = first_data_col + data_col_offset  # 1-based column index
            data_col_letter = get_column_letter(data_col_idx)

            # Build criteria pairs for row filters
            row_criteria_parts: list[str] = []
            for filter_idx, filter_data_col in enumerate(row_filter_cols):
                label_col_letter = get_column_letter(filter_idx + 1)  # A, B, C, ...
                row_criteria_parts.append(
                    f"{data_sheet}!${filter_data_col}:${filter_data_col},"
                    f"${label_col_letter}{row_idx}"
                )

            # Column dimension criterion: relative col, absolute row 2
            col_dim_criterion = (
                f"{data_sheet}!${col_filter_col}:${col_filter_col},"
                f"{data_col_letter}$2"
            )

            all_criteria = ",".join(row_criteria_parts + [col_dim_criterion])
            formula = (
                f"=SUMIFS({data_sheet}!${value_col}:${value_col},{all_criteria})"
            )

            cell = ws.cell(row=row_idx, column=data_col_idx, value=formula)
            apply_normal_style(cell, style)
            if is_alt:
                cell.fill = _ALT_ROW_FILL
            cell.number_format = number_format_data

        # Total column
        if append_total:
            total_col_idx = last_data_col + 1
            first_data_col_letter = get_column_letter(first_data_col)
            last_data_col_letter = get_column_letter(last_data_col)
            total_formula = (
                f"=SUM({first_data_col_letter}{row_idx}:{last_data_col_letter}{row_idx})"
            )
            cell = ws.cell(row=row_idx, column=total_col_idx, value=total_formula)
            apply_normal_style(cell, style)
            if is_alt:
                cell.fill = _ALT_ROW_FILL
            cell.number_format = number_format_data

        # YoY columns (one per consecutive data col pair)
        if append_yoy:
            yoy_start_col = last_data_col + (2 if append_total else 1)
            for yoy_idx in range(n_data_cols - 1):
                prev_col_letter = get_column_letter(first_data_col + yoy_idx)
                cur_col_letter = get_column_letter(first_data_col + yoy_idx + 1)
                yoy_col_idx = yoy_start_col + yoy_idx
                yoy_formula = (
                    f"=IF({prev_col_letter}{row_idx}=0,0,"
                    f"({cur_col_letter}{row_idx}-{prev_col_letter}{row_idx})"
                    f"/ABS({prev_col_letter}{row_idx}))"
                )
                cell = ws.cell(row=row_idx, column=yoy_col_idx, value=yoy_formula)
                apply_normal_style(cell, style)
                if is_alt:
                    cell.fill = _ALT_ROW_FILL
                cell.number_format = number_format_pct

    # Column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{freeze_row + 1}"

    return ws
