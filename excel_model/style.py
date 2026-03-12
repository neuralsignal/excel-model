"""StyleConfig dataclass and openpyxl cell applier functions."""

from dataclasses import dataclass
from pathlib import Path

import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class StyleConfig:
    header_fill_hex: str
    header_font_color: str
    subtotal_fill_hex: str
    total_fill_hex: str
    history_col_fill_hex: str
    section_header_fill_hex: str
    font_name: str
    font_size: int
    number_format_currency: str
    number_format_percent: str
    number_format_integer: str
    number_format_number: str


def load_style_config(path: str) -> StyleConfig:
    """Load StyleConfig from a YAML file. All fields required."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Style config not found: {path}")
    with p.open() as f:
        data = yaml.safe_load(f)
    required = [
        "header_fill_hex",
        "header_font_color",
        "subtotal_fill_hex",
        "total_fill_hex",
        "history_col_fill_hex",
        "section_header_fill_hex",
        "font_name",
        "font_size",
        "number_format_currency",
        "number_format_percent",
        "number_format_integer",
        "number_format_number",
    ]
    missing = [k for k in required if k not in data]
    if missing:
        raise ValueError(f"Style config missing required keys: {missing}")
    return StyleConfig(
        header_fill_hex=data["header_fill_hex"],
        header_font_color=data["header_font_color"],
        subtotal_fill_hex=data["subtotal_fill_hex"],
        total_fill_hex=data["total_fill_hex"],
        history_col_fill_hex=data["history_col_fill_hex"],
        section_header_fill_hex=data["section_header_fill_hex"],
        font_name=data["font_name"],
        font_size=int(data["font_size"]),
        number_format_currency=data["number_format_currency"],
        number_format_percent=data["number_format_percent"],
        number_format_integer=data["number_format_integer"],
        number_format_number=data["number_format_number"],
    )


def _make_fill(hex_color: str) -> PatternFill:
    color = hex_color.lstrip("#")
    return PatternFill(fill_type="solid", fgColor=color)


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(bottom=thin)


def _double_border() -> Border:
    double = Side(style="double")
    return Border(bottom=double)


def _top_border() -> Border:
    thin = Side(style="thin")
    return Border(top=thin)


def apply_header_style(cell, config: StyleConfig) -> None:
    cell.fill = _make_fill(config.header_fill_hex)
    cell.font = Font(
        name=config.font_name,
        size=config.font_size,
        bold=True,
        color=config.header_font_color.lstrip("#"),
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_subtotal_style(cell, config: StyleConfig) -> None:
    cell.fill = _make_fill(config.subtotal_fill_hex)
    cell.font = Font(name=config.font_name, size=config.font_size, bold=True)
    cell.border = _top_border()


def apply_total_style(cell, config: StyleConfig) -> None:
    cell.fill = _make_fill(config.total_fill_hex)
    cell.font = Font(name=config.font_name, size=config.font_size, bold=True)
    cell.border = _double_border()


def apply_section_header_style(cell, config: StyleConfig) -> None:
    cell.fill = _make_fill(config.section_header_fill_hex)
    cell.font = Font(name=config.font_name, size=config.font_size, bold=True, italic=True)


def apply_history_col_style(cell, config: StyleConfig) -> None:
    cell.fill = _make_fill(config.history_col_fill_hex)


def apply_normal_style(cell, config: StyleConfig) -> None:
    cell.font = Font(name=config.font_name, size=config.font_size)


def apply_conditional_formatting(
    ws,
    cell_range: str,
    positive_is_good: bool,
    config: StyleConfig,
) -> None:
    """Apply red/green fill conditional formatting based on cell value sign.

    Green (#C6EFCE) for favorable values, red (#FFC7CE) for unfavorable.
    positive_is_good=True: positive = green, negative = red.
    positive_is_good=False: positive = red, negative = green (e.g. costs).
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as _PatternFill

    green_fill = _PatternFill(fgColor="C6EFCE", fill_type="solid")
    red_fill = _PatternFill(fgColor="FFC7CE", fill_type="solid")

    if positive_is_good:
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule("greaterThan", ["0"], fill=green_fill),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule("lessThan", ["0"], fill=red_fill),
        )
    else:
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule("greaterThan", ["0"], fill=red_fill),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule("lessThan", ["0"], fill=green_fill),
        )


def apply_assumption_sheet_validation(
    ws,
    assumptions: list,
    value_col: int,
    format_col: int,
    start_row: int,
) -> None:
    """Add Excel data validation to the Assumptions sheet.

    - Format column: dropdown list of allowed format types.
    - Value column: 0-1 numeric bounds for percent-type assumptions.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    n = len(assumptions)
    if n == 0:
        return

    end_row = start_row + n - 1
    format_col_letter = get_column_letter(format_col)
    value_col_letter = get_column_letter(value_col)

    # Dropdown on Format column
    format_range = f"{format_col_letter}{start_row}:{format_col_letter}{end_row}"
    dv_format = DataValidation(
        type="list",
        formula1='"number,percent,currency,integer"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid format",
        error="Must be one of: number, percent, currency, integer",
    )
    dv_format.sqref = format_range
    ws.add_data_validation(dv_format)

    # Per-row decimal bounds for percent-type assumptions
    for i, assumption in enumerate(assumptions):
        if getattr(assumption, "format", None) == "percent":
            row = start_row + i
            cell_ref = f"{value_col_letter}{row}"
            dv_pct = DataValidation(
                type="decimal",
                operator="between",
                formula1="0",
                formula2="1",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Out of range",
                error="Percent values must be between 0 and 1",
            )
            dv_pct.sqref = cell_ref
            ws.add_data_validation(dv_pct)


def get_number_format(format_type: str, config: StyleConfig) -> str:
    """Return Excel number format string for the given format_type.

    format_type: 'currency' | 'percent' | 'integer' | 'number'
    """
    mapping = {
        "currency": config.number_format_currency,
        "percent": config.number_format_percent,
        "integer": config.number_format_integer,
        "number": config.number_format_number,
    }
    if format_type not in mapping:
        raise ValueError(f"Unknown format_type: {format_type!r}. Must be one of {list(mapping)}")
    return mapping[format_type]
