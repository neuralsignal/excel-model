"""Tests for excel_writer.py — integration tests that write real xlsx files."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from excel_model.excel_writer import build_workbook
from excel_model.spec import (
    AssumptionDef,
    ColumnGroupDef,
    EntityDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
    ScenarioDef,
)
from excel_model.style import StyleConfig


def make_p_and_l_spec() -> ModelSpec:
    return ModelSpec(
        model_type="p_and_l",
        title="Test P&L",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=2,
        assumptions=(
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth Rate", value=0.10, format="percent", group="Growth"
            ),
            AssumptionDef(name="COGSMargin", label="COGS Margin", value=0.45, format="percent", group="Margins"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="cogs",
                label="  COGS",
                formula_type="pct_of_revenue",
                formula_params={"revenue_key": "revenue", "rate_assumption": "COGSMargin"},
                is_subtotal=False,
                is_total=False,
                section="Cost",
                format="",
            ),
            LineItemDef(
                key="gross_profit",
                label="Gross Profit",
                formula_type="subtraction",
                formula_params={"minuend_key": "revenue", "subtrahend_key": "cogs"},
                is_subtotal=True,
                is_total=False,
                section="Profit",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="Test", date="2026-01-01", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


@pytest.fixture
def p_and_l_spec():
    return make_p_and_l_spec()


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
        alt_row_fill_hex="F2F2F2",
    )


class TestBuildWorkbook:
    def test_creates_file(self, p_and_l_spec, style, tmp_path):
        output = str(tmp_path / "test.xlsx")
        build_workbook(p_and_l_spec, None, output, style)
        assert Path(output).exists()

    def test_has_three_sheets(self, p_and_l_spec, style, tmp_path):
        output = str(tmp_path / "test.xlsx")
        build_workbook(p_and_l_spec, None, output, style)
        wb = load_workbook(output)
        assert "Assumptions" in wb.sheetnames
        assert "Inputs" in wb.sheetnames
        assert "Model" in wb.sheetnames

    def test_named_ranges_present(self, p_and_l_spec, style, tmp_path):
        output = str(tmp_path / "test.xlsx")
        build_workbook(p_and_l_spec, None, output, style)
        wb = load_workbook(output)
        assert "RevenueGrowthRate" in wb.defined_names
        assert "COGSMargin" in wb.defined_names

    def test_assumptions_values_in_cells(self, p_and_l_spec, style, tmp_path):
        output = str(tmp_path / "test.xlsx")
        build_workbook(p_and_l_spec, None, output, style)
        wb = load_workbook(output)
        ws = wb["Assumptions"]
        # Values should appear in column C (col 3)
        values = [ws.cell(row=r, column=3).value for r in range(1, 15)]
        assert 0.10 in values or any(v == 0.10 for v in values)

    def test_model_has_title(self, p_and_l_spec, style, tmp_path):
        output = str(tmp_path / "test.xlsx")
        build_workbook(p_and_l_spec, None, output, style)
        wb = load_workbook(output)
        ws = wb["Model"]
        assert ws["A1"].value == "Test P&L"

    def test_dcf_model(self, style, tmp_path):
        dcf_spec = ModelSpec(
            model_type="dcf",
            title="Test DCF",
            currency="CHF",
            granularity="annual",
            start_period="2025",
            n_periods=5,
            n_history_periods=0,
            assumptions=(AssumptionDef(name="WACC", label="WACC", value=0.10, format="percent", group="Valuation"),),
            line_items=(
                LineItemDef(
                    key="revenue",
                    label="Revenue",
                    formula_type="growth_projected",
                    formula_params={"growth_assumption": "WACC"},
                    is_subtotal=False,
                    is_total=False,
                    section="Income",
                    format="",
                ),
            ),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(),
            column_groups=(),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(),
            drivers=(),
        )
        output = str(tmp_path / "dcf.xlsx")
        build_workbook(dcf_spec, None, output, style)
        wb = load_workbook(output)
        assert "Model" in wb.sheetnames
        assert "WACC" in wb.defined_names

    def test_scenario_model(self, style, tmp_path):
        scenario_spec = ModelSpec(
            model_type="scenario",
            title="Scenario Test",
            currency="CHF",
            granularity="annual",
            start_period="2025",
            n_periods=3,
            n_history_periods=0,
            assumptions=(
                AssumptionDef(
                    name="RevenueGrowthRate", label="Rev Growth", value=0.10, format="percent", group="Growth"
                ),
            ),
            line_items=(
                LineItemDef(
                    key="revenue",
                    label="Revenue",
                    formula_type="growth_projected",
                    formula_params={"growth_assumption": "RevenueGrowthRate"},
                    is_subtotal=False,
                    is_total=False,
                    section="Revenue",
                    format="",
                ),
            ),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(
                ScenarioDef(name="base", label="Base Case", assumption_overrides={}, driver_overrides={}),
                ScenarioDef(
                    name="bull",
                    label="Bull Case",
                    assumption_overrides={"RevenueGrowthRate": 0.20},
                    driver_overrides={},
                ),
            ),
            column_groups=(),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(),
            drivers=(),
        )
        output = str(tmp_path / "scenario.xlsx")
        build_workbook(scenario_spec, None, output, style)
        wb = load_workbook(output)
        assert "Model" in wb.sheetnames
        assert "BullRevenueGrowthRate" in wb.defined_names

    def test_comparison_model(self, style, tmp_path):
        comparison_spec = ModelSpec(
            model_type="comparison",
            title="Comparison Test",
            currency="CHF",
            granularity="auto",
            start_period="2025",
            n_periods=0,
            n_history_periods=0,
            assumptions=(),
            line_items=(
                LineItemDef(
                    key="revenue",
                    label="Revenue",
                    formula_type="constant",
                    formula_params={"value": 0},
                    is_subtotal=False,
                    is_total=False,
                    section="Revenue",
                    format="",
                ),
            ),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(),
            column_groups=(),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(
                EntityDef(key="entity_a", label="Entity A"),
                EntityDef(key="entity_b", label="Entity B"),
            ),
            drivers=(),
        )
        output = str(tmp_path / "comparison.xlsx")
        build_workbook(comparison_spec, None, output, style)
        wb = load_workbook(output)
        assert Path(output).exists()
        assert "Model" in wb.sheetnames

    def test_budget_vs_actuals_model(self, style, tmp_path):
        bva_spec = ModelSpec(
            model_type="budget_vs_actuals",
            title="BvA Test",
            currency="CHF",
            granularity="monthly",
            start_period="2025-01",
            n_periods=3,
            n_history_periods=0,
            assumptions=(
                AssumptionDef(
                    name="RevenueGrowthRate", label="Rev Growth", value=0.08, format="percent", group="Budget"
                ),
            ),
            line_items=(
                LineItemDef(
                    key="revenue_plan",
                    label="Revenue (Plan)",
                    formula_type="constant",
                    formula_params={"value": 1000},
                    is_subtotal=False,
                    is_total=False,
                    section="Revenue",
                    format="",
                ),
                LineItemDef(
                    key="revenue_actual",
                    label="Revenue (Actual)",
                    formula_type="constant",
                    formula_params={"value": 1050},
                    is_subtotal=False,
                    is_total=False,
                    section="Revenue",
                    format="",
                ),
                LineItemDef(
                    key="revenue",
                    label="Revenue Variance",
                    formula_type="variance",
                    formula_params={"plan_key": "revenue_plan", "actual_key": "revenue_actual"},
                    is_subtotal=False,
                    is_total=False,
                    section="Revenue",
                    format="",
                ),
            ),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(),
            column_groups=(
                ColumnGroupDef(key="plan", label="Plan", color_hex="D6E4F0"),
                ColumnGroupDef(key="actual", label="Actual", color_hex="FDFEFE"),
                ColumnGroupDef(key="variance", label="Variance", color_hex="FEF9E7"),
            ),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(),
            drivers=(),
        )
        output = str(tmp_path / "bva.xlsx")
        build_workbook(bva_spec, None, output, style)
        wb = load_workbook(output)
        assert Path(output).exists()
        assert "Model" in wb.sheetnames

    def test_custom_model_dispatches_to_p_and_l(self, style, tmp_path):
        custom_spec = make_p_and_l_spec()
        custom_spec = ModelSpec(
            model_type="custom",
            title=custom_spec.title,
            currency=custom_spec.currency,
            granularity=custom_spec.granularity,
            start_period=custom_spec.start_period,
            n_periods=custom_spec.n_periods,
            n_history_periods=custom_spec.n_history_periods,
            assumptions=custom_spec.assumptions,
            line_items=custom_spec.line_items,
            metadata=custom_spec.metadata,
            scenarios=custom_spec.scenarios,
            column_groups=custom_spec.column_groups,
            inputs=custom_spec.inputs,
            entities=custom_spec.entities,
            drivers=custom_spec.drivers,
        )
        output = str(tmp_path / "custom.xlsx")
        build_workbook(custom_spec, None, output, style)
        wb = load_workbook(output)
        assert Path(output).exists()
        assert "Model" in wb.sheetnames

    def test_unknown_model_type_raises(self, style, tmp_path):
        bad_spec = ModelSpec(
            model_type="nonexistent",
            title="Bad",
            currency="CHF",
            granularity="annual",
            start_period="2025",
            n_periods=1,
            n_history_periods=0,
            assumptions=(),
            line_items=(),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(),
            column_groups=(),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(),
            drivers=(),
        )
        output = str(tmp_path / "bad.xlsx")
        with pytest.raises(ValueError, match="Unknown model_type"):
            build_workbook(bad_spec, None, output, style)
