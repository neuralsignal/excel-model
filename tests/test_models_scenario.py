"""Tests for Scenario model builder."""

import pytest
from openpyxl import Workbook

from excel_model.models.scenario import build_scenario
from excel_model.spec import (
    AssumptionDef,
    DriverDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
    ScenarioDef,
)
from excel_model.style import StyleConfig
from excel_model.time_engine import generate_periods


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
        alt_row_fill_hex="F2F2F2",
    )


@pytest.fixture
def scenario_spec():
    return ModelSpec(
        model_type="scenario",
        title="Scenario Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="RevenueGrowthRate", label="Rev Growth", value=0.10, format="percent", group="Growth"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(
            ScenarioDef(name="base", label="Base Case", assumption_overrides={}, driver_overrides={}),
            ScenarioDef(
                name="bull", label="Bull Case", assumption_overrides={"RevenueGrowthRate": 0.20}, driver_overrides={}
            ),
            ScenarioDef(
                name="bear", label="Bear Case", assumption_overrides={"RevenueGrowthRate": 0.02}, driver_overrides={}
            ),
        ),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def test_scenario_creates_sheets(scenario_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_scenario(wb, scenario_spec, None, style, periods)
    assert "Assumptions" in wb.sheetnames
    assert "Inputs" in wb.sheetnames
    assert "Model" in wb.sheetnames


def test_scenario_registers_per_scenario_named_ranges(scenario_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_scenario(wb, scenario_spec, None, style, periods)
    # Should have Base, Bull, Bear named ranges
    assert "BaseRevenueGrowthRate" in wb.defined_names
    assert "BullRevenueGrowthRate" in wb.defined_names
    assert "BearRevenueGrowthRate" in wb.defined_names


def test_scenario_bull_override_value(scenario_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_scenario(wb, scenario_spec, None, style, periods)

    ws = wb["Assumptions"]
    # Find the cell with BullRevenueGrowthRate value (should be 0.20)
    found_value = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 0.20:
                found_value = True
    assert found_value, "Bull case override value 0.20 not found in Assumptions sheet"


def test_scenario_custom_formulas_use_prefixed_names(style):
    """Custom formulas with assumption names must get scenario-prefixed names."""
    spec = ModelSpec(
        model_type="scenario",
        title="Custom Formula Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=2,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="PatientCount", label="Patient Count", value=500, format="integer", group="Volume"),
            AssumptionDef(name="PricePerPatient", label="Price/Patient", value=100, format="currency", group="Pricing"),
        ),
        line_items=(
            LineItemDef(
                key="total_revenue",
                label="Total Revenue",
                formula_type="custom",
                formula_params={"formula": "=PatientCount*PricePerPatient"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(
            ScenarioDef(name="standard", label="Standard", assumption_overrides={}, driver_overrides={}),
            ScenarioDef(
                name="premium", label="Premium", assumption_overrides={"PricePerPatient": 200}, driver_overrides={}
            ),
        ),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 2, 0, "annual")
    build_scenario(wb, spec, None, style, periods)

    ws = wb["Model"]
    # Scan all cells for formulas — they must contain prefixed names
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.value)

    assert formulas, "Expected at least one formula in Model sheet"
    # Every formula referencing PatientCount should have a prefix
    for f in formulas:
        assert "PatientCount" not in f or "StandardPatientCount" in f or "PremiumPatientCount" in f, (
            f"Formula contains unprefixed PatientCount: {f}"
        )
    # At least one Standard and one Premium prefixed name
    assert any("StandardPatientCount" in f for f in formulas)
    assert any("PremiumPatientCount" in f for f in formulas)


def test_scenario_model_has_scenario_labels(scenario_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_scenario(wb, scenario_spec, None, style, periods)
    ws = wb["Model"]
    # Row 3 should have scenario labels
    sub_labels = [ws.cell(row=3, column=c).value for c in range(2, 12)]
    assert "Base Case" in sub_labels
    assert "Bull Case" in sub_labels
    assert "Bear Case" in sub_labels


# ─── Tests for the new drivers concept ───


def _make_driver_spec(style_fixture_unused=None):
    """Spec with drivers separated from assumptions."""
    return ModelSpec(
        model_type="scenario",
        title="Driver Test",
        currency="EUR",
        granularity="annual",
        start_period="2026",
        n_periods=1,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(
                name="CROPerPatient", label="CRO Per Patient", value=2992, format="currency", group="Benchmarks"
            ),
        ),
        drivers=(
            DriverDef(name="PatientCount", label="Patient Count", value=2500, format="integer", group="Volume"),
            DriverDef(name="PerPatientPrice", label="Price/Patient", value=40, format="currency", group="Pricing"),
        ),
        line_items=(
            LineItemDef(
                key="data_revenue",
                label="Data Revenue",
                formula_type="custom",
                formula_params={"formula": "=PatientCount*PerPatientPrice"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="vs_cro",
                label="vs CRO",
                formula_type="custom",
                formula_params={"formula": "=${col_letter}${data_revenue_row}/PatientCount/CROPerPatient"},
                is_subtotal=False,
                is_total=False,
                section="Positioning",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(
            ScenarioDef(name="standard", label="Standard", assumption_overrides={}, driver_overrides={}),
            ScenarioDef(
                name="premium", label="Premium", assumption_overrides={}, driver_overrides={"PerPatientPrice": 180}
            ),
        ),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
    )


class TestScenarioWithDrivers:
    def test_creates_four_sheets(self, style):
        spec = _make_driver_spec()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2026", 1, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        assert "Assumptions" in wb.sheetnames
        assert "Drivers" in wb.sheetnames
        assert "Inputs" in wb.sheetnames
        assert "Model" in wb.sheetnames

    def test_assumptions_have_bare_named_ranges(self, style):
        spec = _make_driver_spec()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2026", 1, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        # CROPerPatient should be bare (no prefix)
        assert "CROPerPatient" in wb.defined_names
        # Should NOT have prefixed versions
        assert "StandardCROPerPatient" not in wb.defined_names
        assert "PremiumCROPerPatient" not in wb.defined_names

    def test_drivers_have_prefixed_named_ranges(self, style):
        spec = _make_driver_spec()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2026", 1, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        assert "StandardPatientCount" in wb.defined_names
        assert "PremiumPatientCount" in wb.defined_names
        assert "StandardPerPatientPrice" in wb.defined_names
        assert "PremiumPerPatientPrice" in wb.defined_names

    def test_driver_overrides_applied(self, style):
        spec = _make_driver_spec()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2026", 1, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        ws = wb["Drivers"]
        # Find Premium PerPatientPrice value (should be 180)
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 180:
                    found = True
        assert found, "Premium driver_overrides value 180 not found in Drivers sheet"

    def test_model_formulas_mix_bare_and_prefixed(self, style):
        spec = _make_driver_spec()
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2026", 1, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        ws = wb["Model"]
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.value)

        assert formulas, "Expected formulas in Model sheet"
        # CROPerPatient should appear bare (not prefixed)
        cro_formulas = [f for f in formulas if "CROPerPatient" in f]
        for f in cro_formulas:
            assert "StandardCROPerPatient" not in f, f"CROPerPatient should be bare: {f}"
            assert "PremiumCROPerPatient" not in f, f"CROPerPatient should be bare: {f}"

        # Driver names should appear prefixed
        assert any("StandardPatientCount" in f for f in formulas)
        assert any("PremiumPatientCount" in f for f in formulas)

    def test_backward_compat_no_drivers(self, style):
        """Spec without drivers produces identical output to legacy mode."""
        spec = ModelSpec(
            model_type="scenario",
            title="Legacy Test",
            currency="CHF",
            granularity="annual",
            start_period="2025",
            n_periods=2,
            n_history_periods=0,
            assumptions=(AssumptionDef(name="Growth", label="Growth", value=0.10, format="percent", group="G"),),
            drivers=(),  # empty = legacy
            line_items=(
                LineItemDef(
                    key="rev",
                    label="Revenue",
                    formula_type="growth_projected",
                    formula_params={"growth_assumption": "Growth"},
                    is_subtotal=False,
                    is_total=False,
                    section="Rev",
                    format="",
                ),
            ),
            metadata=MetadataDef(preparer="", date="", version="1.0"),
            scenarios=(
                ScenarioDef(name="base", label="Base", assumption_overrides={}, driver_overrides={}),
                ScenarioDef(name="bull", label="Bull", assumption_overrides={"Growth": 0.20}, driver_overrides={}),
            ),
            column_groups=(),
            inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
            entities=(),
        )
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        periods = generate_periods("2025", 2, 0, "annual")
        build_scenario(wb, spec, None, style, periods)
        # Legacy: no Drivers sheet
        assert "Drivers" not in wb.sheetnames
        assert "Assumptions" in wb.sheetnames
        # All assumptions prefixed per scenario
        assert "BaseGrowth" in wb.defined_names
        assert "BullGrowth" in wb.defined_names


# ─── Tests for uncovered code paths ───


def test_scenario_all_history_periods(style):
    """Lines 130-131: all periods are historical, so no projection columns."""
    spec = ModelSpec(
        model_type="scenario",
        title="All History",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=0,
        n_history_periods=2,
        assumptions=(AssumptionDef(name="Growth", label="Growth", value=0.10, format="percent", group="G"),),
        drivers=(),
        line_items=(
            LineItemDef(
                key="rev",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(ScenarioDef(name="base", label="Base", assumption_overrides={}, driver_overrides={}),),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 0, 2, "annual")
    build_scenario(wb, spec, None, style, periods)
    assert "Model" in wb.sheetnames
    ws = wb["Model"]
    # Data cells should contain the constant value
    assert ws.cell(row=4, column=2).value == 100


def test_scenario_with_subtotal_and_total_line_items(style):
    """Lines 191/193/238/240: subtotal and total styling on label and data cells."""
    spec = ModelSpec(
        model_type="scenario",
        title="Subtotal Total Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=1,
        n_history_periods=0,
        assumptions=(AssumptionDef(name="Growth", label="Growth", value=0.10, format="percent", group="G"),),
        drivers=(),
        line_items=(
            LineItemDef(
                key="rev",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="subtot",
                label="Subtotal",
                formula_type="sum_of_rows",
                formula_params={"addend_keys": ["rev"]},
                is_subtotal=True,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="total",
                label="Grand Total",
                formula_type="sum_of_rows",
                formula_params={"addend_keys": ["rev"]},
                is_subtotal=False,
                is_total=True,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(ScenarioDef(name="base", label="Base", assumption_overrides={}, driver_overrides={}),),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 1, 0, "annual")
    build_scenario(wb, spec, None, style, periods)
    ws = wb["Model"]
    # Subtotal label cell (row 5) should have bold font from subtotal style
    subtotal_label = ws.cell(row=5, column=1)
    assert subtotal_label.value == "Subtotal"
    assert subtotal_label.font.bold is True
    # Total label cell (row 6) should have bold font from total style
    total_label = ws.cell(row=6, column=1)
    assert total_label.value == "Grand Total"
    assert total_label.font.bold is True
    # Data cells also get styled
    subtotal_data = ws.cell(row=5, column=2)
    assert subtotal_data.font.bold is True
    total_data = ws.cell(row=6, column=2)
    assert total_data.font.bold is True


def test_scenario_with_input_ref_formula(style):
    """Line 206: input_ref formula type sets line_item_key in params."""
    spec = ModelSpec(
        model_type="scenario",
        title="Input Ref Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=2,
        n_history_periods=0,
        assumptions=(AssumptionDef(name="Growth", label="Growth", value=0.05, format="percent", group="G"),),
        drivers=(),
        line_items=(
            LineItemDef(
                key="rev",
                label="Revenue",
                formula_type="input_ref",
                formula_params={"projected_type": "growth_projected", "growth_assumption": "Growth"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(ScenarioDef(name="base", label="Base", assumption_overrides={}, driver_overrides={}),),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 2, 0, "annual")
    build_scenario(wb, spec, None, style, periods)
    ws = wb["Model"]
    # input_ref with n_history=0 delegates to projected_type, so should produce formulas
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.value)
    assert formulas, "Expected formulas from input_ref projected delegation"


def test_scenario_variance_conditional_formatting(style):
    """Lines 244-248: variance with positive_is_good triggers conditional formatting."""
    spec = ModelSpec(
        model_type="scenario",
        title="Variance CF Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=1,
        n_history_periods=0,
        assumptions=(AssumptionDef(name="Growth", label="Growth", value=0.10, format="percent", group="G"),),
        drivers=(),
        line_items=(
            LineItemDef(
                key="plan",
                label="Plan",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="actual",
                label="Actual",
                formula_type="constant",
                formula_params={"value": 120},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="var",
                label="Variance",
                formula_type="variance",
                formula_params={"plan_key": "plan", "actual_key": "actual", "positive_is_good": True},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="var_pct",
                label="Variance %",
                formula_type="variance_pct",
                formula_params={"plan_key": "plan", "actual_key": "actual", "positive_is_good": False},
                is_subtotal=False,
                is_total=False,
                section="",
                format="percent",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(ScenarioDef(name="base", label="Base", assumption_overrides={}, driver_overrides={}),),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 1, 0, "annual")
    build_scenario(wb, spec, None, style, periods)
    ws = wb["Model"]
    # Variance row (row 6) should have conditional formatting rules
    cf_rules = ws.conditional_formatting
    assert len(cf_rules) >= 2, f"Expected at least 2 conditional formatting rules, got {len(cf_rules)}"
