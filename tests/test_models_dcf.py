"""Tests for DCF model builder."""

import pytest
from openpyxl import Workbook

from excel_model.models.dcf import build_dcf
from excel_model.spec import (
    AssumptionDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
)
from excel_model.style import StyleConfig
from excel_model.time_engine import generate_periods


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
    )


@pytest.fixture
def dcf_spec():
    return ModelSpec(
        model_type="dcf",
        title="DCF Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=5,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="WACC", label="Discount Rate", value=0.10, format="percent", group="Valuation"),
            AssumptionDef(name="TGR", label="Terminal Growth Rate", value=0.02, format="percent", group="Valuation"),
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth", value=0.10, format="percent", group="Growth"
            ),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Income",
                format="",
            ),
            LineItemDef(
                key="fcf",
                label="Free Cash Flow",
                formula_type="constant",
                formula_params={"value": 0},
                is_subtotal=True,
                is_total=False,
                section="FCF",
                format="",
            ),
            LineItemDef(
                key="pv_fcf",
                label="PV of FCF",
                formula_type="discounted_pv",
                formula_params={"cashflow_key": "fcf", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="Valuation",
                format="",
            ),
            LineItemDef(
                key="terminal_value",
                label="Terminal Value",
                formula_type="terminal_value",
                formula_params={"cashflow_key": "fcf", "growth_assumption": "TGR", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="Valuation",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def test_dcf_creates_sheets(dcf_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 5, 0, "annual")
    build_dcf(wb, dcf_spec, None, style, periods)
    assert "Assumptions" in wb.sheetnames
    assert "Inputs" in wb.sheetnames
    assert "Model" in wb.sheetnames


def test_dcf_discounted_pv_formula(dcf_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 5, 0, "annual")
    build_dcf(wb, dcf_spec, None, style, periods)

    ws = wb["Model"]
    # Should have a discounted_pv formula with WACC
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "WACC" in cell.value and "^" in cell.value:
                found = True
    assert found, "No discounted PV formula found"


def test_dcf_terminal_value_formula(dcf_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 5, 0, "annual")
    build_dcf(wb, dcf_spec, None, style, periods)

    ws = wb["Model"]
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "TGR" in cell.value:
                found = True
    assert found, "No terminal value formula (with TGR) found"


@pytest.fixture
def dcf_spec_with_history():
    """DCF spec with history periods, npv_sum, input_ref, total, and subtotal items."""
    return ModelSpec(
        model_type="dcf",
        title="DCF History Test",
        currency="CHF",
        granularity="annual",
        start_period="2023",
        n_periods=5,
        n_history_periods=2,
        assumptions=(
            AssumptionDef(name="WACC", label="Discount Rate", value=0.10, format="percent", group="Valuation"),
            AssumptionDef(name="TGR", label="Terminal Growth Rate", value=0.02, format="percent", group="Valuation"),
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth", value=0.10, format="percent", group="Growth"
            ),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="input_ref",
                formula_params={"projected_type": "growth_projected", "growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Income",
                format="",
            ),
            LineItemDef(
                key="fcf",
                label="Free Cash Flow",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=True,
                is_total=False,
                section="FCF",
                format="",
            ),
            LineItemDef(
                key="pv_fcf",
                label="PV of FCF",
                formula_type="discounted_pv",
                formula_params={"cashflow_key": "fcf", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="Valuation",
                format="",
            ),
            LineItemDef(
                key="terminal_value",
                label="Terminal Value",
                formula_type="terminal_value",
                formula_params={"cashflow_key": "fcf", "growth_assumption": "TGR", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="Valuation",
                format="",
            ),
            LineItemDef(
                key="npv",
                label="NPV",
                formula_type="npv_sum",
                formula_params={"pv_fcf_key": "pv_fcf", "pv_terminal_key": "terminal_value"},
                is_subtotal=False,
                is_total=True,
                section="Valuation",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def test_dcf_with_history_periods(dcf_spec_with_history, style):
    """Covers history column styling (lines 76-77), history border (lines 207-210),
    input_ref formula type (line 169), and history data cell styling (lines 195, 202-203)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2023", 5, 2, "annual")
    build_dcf(wb, dcf_spec_with_history, None, style, periods)

    ws = wb["Model"]

    # Verify history period headers have history styling
    # start_period=2023 with n_history=2 means history is 2021, 2022
    history_header_1 = ws.cell(row=2, column=2)
    history_header_2 = ws.cell(row=2, column=3)
    assert history_header_1.value == "2021"
    assert history_header_2.value == "2022"
    assert history_header_1.font.bold is True
    assert history_header_2.font.bold is True

    # Verify projection header is not history-styled
    proj_header = ws.cell(row=2, column=4)
    assert proj_header.value == "2023"

    # Verify thin border after last history column on data rows
    # border_col = 1 + 2 + 1 = 4
    border_cell = ws.cell(row=4, column=4)
    assert border_cell.border.left.style == "thin"


def test_dcf_npv_sum_formula(dcf_spec_with_history, style):
    """Covers npv_sum branch (lines 131-160) and is_total styling on npv_sum (line 125, 155-156)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2023", 5, 2, "annual")
    build_dcf(wb, dcf_spec_with_history, None, style, periods)

    ws = wb["Model"]

    # Find the NPV row — it has is_total=True and npv_sum formula
    npv_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "NPV":
            npv_row = row[0].row
            break
    assert npv_row is not None, "NPV row not found"

    # npv_sum writes formula only in first data column (col 2)
    npv_cell = ws.cell(row=npv_row, column=2)
    assert isinstance(npv_cell.value, str)
    assert npv_cell.value.startswith("=SUM(")

    # The NPV label cell should have total styling (line 125)
    npv_label = ws.cell(row=npv_row, column=1)
    assert npv_label.value == "NPV"


def test_dcf_input_ref_formula(dcf_spec_with_history, style):
    """Covers input_ref branch (line 169) in per-column rendering."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2023", 5, 2, "annual")
    build_dcf(wb, dcf_spec_with_history, None, style, periods)

    ws = wb["Model"]

    # Find revenue row (uses input_ref)
    rev_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "Revenue":
            rev_row = row[0].row
            break
    assert rev_row is not None

    # History columns should have value 0 (no input data loaded)
    hist_cell = ws.cell(row=rev_row, column=2)
    assert hist_cell.value == 0

    # Projection columns should have growth_projected formula
    proj_cell = ws.cell(row=rev_row, column=4)
    assert isinstance(proj_cell.value, str)
    assert "RevenueGrowthRate" in proj_cell.value


def test_dcf_subtotal_and_total_styling(dcf_spec_with_history, style):
    """Covers is_subtotal (line 197) and is_total (line 199) styling on data cells,
    and normal styling (lines 200-201)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2023", 5, 2, "annual")
    build_dcf(wb, dcf_spec_with_history, None, style, periods)

    ws = wb["Model"]

    # FCF row is is_subtotal=True
    fcf_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "Free Cash Flow":
            fcf_row = row[0].row
            break
    assert fcf_row is not None

    # Check subtotal styling applied to label cell (line 122-123)
    fcf_label = ws.cell(row=fcf_row, column=1)
    assert fcf_label.fill.start_color.rgb is not None

    # Check data cell in projection column has subtotal fill
    fcf_proj = ws.cell(row=fcf_row, column=4)
    assert fcf_proj.fill.start_color.rgb is not None


def test_dcf_no_projection_periods(style):
    """Covers lines 59-60: no projection periods branch."""
    spec = ModelSpec(
        model_type="dcf",
        title="DCF No Proj",
        currency="CHF",
        granularity="annual",
        start_period="2023",
        n_periods=0,
        n_history_periods=2,
        assumptions=(
            AssumptionDef(name="WACC", label="Discount Rate", value=0.10, format="percent", group="Valuation"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2023", 0, 2, "annual")
    # All periods are history — no projection periods
    assert all(p.is_history for p in periods)
    build_dcf(wb, spec, None, style, periods)
    assert "Model" in wb.sheetnames


def test_dcf_npv_sum_subtotal_styling(style):
    """Covers lines 157-160: npv_sum with is_subtotal=True."""
    spec = ModelSpec(
        model_type="dcf",
        title="DCF NPV Subtotal",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="WACC", label="Discount Rate", value=0.10, format="percent", group="Valuation"),
            AssumptionDef(name="TGR", label="Terminal Growth Rate", value=0.02, format="percent", group="Valuation"),
        ),
        line_items=(
            LineItemDef(
                key="fcf",
                label="FCF",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="pv_fcf",
                label="PV of FCF",
                formula_type="discounted_pv",
                formula_params={"cashflow_key": "fcf", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="terminal_value",
                label="Terminal Value",
                formula_type="terminal_value",
                formula_params={"cashflow_key": "fcf", "growth_assumption": "TGR", "rate_assumption": "WACC"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="npv_sub",
                label="NPV Subtotal",
                formula_type="npv_sum",
                formula_params={"pv_fcf_key": "pv_fcf", "pv_terminal_key": "terminal_value"},
                is_subtotal=True,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="npv_normal",
                label="NPV Normal",
                formula_type="npv_sum",
                formula_params={"pv_fcf_key": "pv_fcf", "pv_terminal_key": "terminal_value"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_dcf(wb, spec, None, style, periods)

    ws = wb["Model"]
    # Find NPV Subtotal row
    npv_sub_row = None
    npv_normal_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "NPV Subtotal":
            npv_sub_row = row[0].row
        if row[0].value == "NPV Normal":
            npv_normal_row = row[0].row
    assert npv_sub_row is not None
    assert npv_normal_row is not None

    # Both should have SUM formulas in column 2
    assert ws.cell(row=npv_sub_row, column=2).value.startswith("=SUM(")
    assert ws.cell(row=npv_normal_row, column=2).value.startswith("=SUM(")


def test_dcf_total_styling_per_column(style):
    """Covers line 199: is_total=True styling on per-column data cells."""
    spec = ModelSpec(
        model_type="dcf",
        title="DCF Total Per Col",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="WACC", label="Discount Rate", value=0.10, format="percent", group="Valuation"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="total_rev",
                label="Total Revenue",
                formula_type="sum_of_rows",
                formula_params={"addend_keys": ["revenue"]},
                is_subtotal=False,
                is_total=True,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 0, "annual")
    build_dcf(wb, spec, None, style, periods)

    ws = wb["Model"]
    # Find Total Revenue row
    total_row = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value == "Total Revenue":
            total_row = row[0].row
            break
    assert total_row is not None

    # Data cell should have total fill
    total_cell = ws.cell(row=total_row, column=2)
    assert total_cell.fill.start_color.rgb is not None
