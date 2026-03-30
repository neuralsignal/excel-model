"""Tests for P&L model builder."""

import pytest
from openpyxl import Workbook

from excel_model.models.p_and_l import build_p_and_l
from excel_model.spec import (
    AssumptionDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
)
from excel_model.style import StyleConfig
from excel_model.time_engine import generate_periods


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
    )


@pytest.fixture
def p_and_l_spec():
    return ModelSpec(
        model_type="p_and_l",
        title="P&L Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=2,
        assumptions=(
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth Rate", value=0.10, format="percent", group="Growth"
            ),
            AssumptionDef(name="COGSMargin", label="COGS Margin", value=0.45, format="percent", group="Margins"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="cogs",
                label="  COGS",
                formula_type="pct_of_revenue",
                formula_params={"revenue_key": "revenue", "rate_assumption": "COGSMargin"},
                is_subtotal=False,
                is_total=False,
                section="Cost",
                format="",
            ),
            LineItemDef(
                key="gross_profit",
                label="Gross Profit",
                formula_type="subtraction",
                formula_params={"minuend_key": "revenue", "subtrahend_key": "cogs"},
                is_subtotal=True,
                is_total=False,
                section="Profit",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def test_p_and_l_creates_three_sheets(p_and_l_spec, style, tmp_path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 2, "annual")
    build_p_and_l(wb, p_and_l_spec, None, style, periods)
    assert "Assumptions" in wb.sheetnames
    assert "Inputs" in wb.sheetnames
    assert "Model" in wb.sheetnames


def test_p_and_l_model_sheet_has_periods(p_and_l_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 2, "annual")
    build_p_and_l(wb, p_and_l_spec, None, style, periods)

    ws = wb["Model"]
    # Row 2 should have "Line Item" in A2 and period labels in subsequent cols
    assert ws["A2"].value == "Line Item"
    period_labels = [ws.cell(row=2, column=c).value for c in range(2, 7)]
    assert "2023" in period_labels
    assert "2025" in period_labels
    assert "2027" in period_labels


def test_p_and_l_assumptions_have_named_ranges(p_and_l_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 2, "annual")
    build_p_and_l(wb, p_and_l_spec, None, style, periods)
    assert "RevenueGrowthRate" in wb.defined_names
    assert "COGSMargin" in wb.defined_names


def test_p_and_l_formulas_written(p_and_l_spec, style, tmp_path):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 3, 2, "annual")
    build_p_and_l(wb, p_and_l_spec, None, style, periods)

    ws = wb["Model"]
    # Find a formula cell that should reference RevenueGrowthRate
    found_formula = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "RevenueGrowthRate" in cell.value:
                found_formula = True
                break
    assert found_formula, "No formula containing RevenueGrowthRate found in Model sheet"


def test_p_and_l_all_history_periods(style):
    """Lines 66-67: all-history P&L with n_periods=0."""
    spec = ModelSpec(
        model_type="p_and_l",
        title="All History P&L",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=0,
        n_history_periods=3,
        assumptions=(
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth Rate", value=0.10, format="percent", group="Growth"
            ),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 0, 3, "annual")
    build_p_and_l(wb, spec, None, style, periods)
    ws = wb["Model"]
    # All periods should be history
    assert all(p.is_history for p in periods)
    # Model sheet should still be created
    assert ws is not None


def test_p_and_l_with_total_line_item(style):
    """Lines 130/170: line item with is_total=True triggers apply_total_style."""
    spec = ModelSpec(
        model_type="p_and_l",
        title="P&L With Total",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=2,
        n_history_periods=1,
        assumptions=(
            AssumptionDef(
                name="RevenueGrowthRate", label="Revenue Growth Rate", value=0.10, format="percent", group="Growth"
            ),
            AssumptionDef(name="COGSMargin", label="COGS Margin", value=0.45, format="percent", group="Margins"),
        ),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="growth_projected",
                formula_params={"growth_assumption": "RevenueGrowthRate"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="cogs",
                label="  COGS",
                formula_type="pct_of_revenue",
                formula_params={"revenue_key": "revenue", "rate_assumption": "COGSMargin"},
                is_subtotal=False,
                is_total=False,
                section="Cost",
                format="",
            ),
            LineItemDef(
                key="net_income",
                label="Net Income",
                formula_type="subtraction",
                formula_params={"minuend_key": "revenue", "subtrahend_key": "cogs"},
                is_subtotal=False,
                is_total=True,
                section="Total",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 2, 1, "annual")
    build_p_and_l(wb, spec, None, style, periods)
    ws = wb["Model"]
    # Find the "Net Income" label cell and verify total style applied
    total_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == "Net Income":
                total_row = cell.row
                break
    assert total_row is not None, "Net Income row not found"
    label_cell = ws.cell(row=total_row, column=1)
    # Total style sets a bold font and a fill color
    assert label_cell.font.bold is True


def test_p_and_l_with_input_ref_formula(style):
    """Line 139: input_ref formula type sets line_item_key param."""
    spec = ModelSpec(
        model_type="p_and_l",
        title="P&L Input Ref",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=2,
        n_history_periods=1,
        assumptions=(),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="input_ref",
                formula_params={"projected_type": "constant", "value": 1000},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025", 2, 1, "annual")
    build_p_and_l(wb, spec, None, style, periods)
    ws = wb["Model"]
    # History column (col 2) should have an input_ref value (0 since no input data)
    history_cell = ws.cell(row=3, column=2)
    assert history_cell.value == 0
    # Projection columns should have the constant value
    proj_cell = ws.cell(row=3, column=3)
    assert proj_cell.value == 1000
