"""Tests for Comparison model builder."""

import pytest
from openpyxl import Workbook

from excel_model.models.comparison import _entity_col_index, build_comparison
from excel_model.spec import (
    EntityDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
)
from excel_model.style import StyleConfig


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
    )


@pytest.fixture
def comparison_spec():
    return ModelSpec(
        model_type="comparison",
        title="Competitive Comparison",
        currency="EUR",
        granularity="auto",
        start_period="2025",
        n_periods=0,
        n_history_periods=0,
        assumptions=(),
        line_items=(
            LineItemDef(
                key="revenue",
                label="Revenue (EUR M)",
                formula_type="constant",
                formula_params={"value": 0},
                is_subtotal=False,
                is_total=False,
                section="Financials",
                format="",
            ),
            LineItemDef(
                key="ebitda",
                label="EBITDA (EUR M)",
                formula_type="constant",
                formula_params={"value": 0},
                is_subtotal=False,
                is_total=False,
                section="Financials",
                format="",
            ),
            LineItemDef(
                key="ebitda_margin",
                label="EBITDA Margin",
                formula_type="ratio",
                formula_params={"numerator_key": "ebitda", "denominator_key": "revenue"},
                is_subtotal=False,
                is_total=False,
                section="Efficiency",
                format="",
            ),
            LineItemDef(
                key="revenue_rank",
                label="Revenue Rank",
                formula_type="rank",
                formula_params={"value_key": "revenue"},
                is_subtotal=False,
                is_total=False,
                section="Rankings",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(
            EntityDef(key="company_a", label="Company A"),
            EntityDef(key="company_b", label="Company B"),
            EntityDef(key="company_c", label="Company C"),
        ),
        drivers=(),
    )


def test_comparison_creates_sheets(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    assert "Assumptions" in wb.sheetnames
    assert "Model" in wb.sheetnames


def test_comparison_entity_headers(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    ws = wb["Model"]
    # Row 2 should have "Metric" in A2 and entity labels
    assert ws["A2"].value == "Metric"
    labels = [ws.cell(row=2, column=c).value for c in range(2, 5)]
    assert "Company A" in labels
    assert "Company B" in labels
    assert "Company C" in labels


def test_comparison_has_section_headers(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    ws = wb["Model"]
    # Look for section headers
    all_values = [ws.cell(row=r, column=1).value for r in range(3, 15)]
    assert "Financials" in all_values
    assert "Efficiency" in all_values
    assert "Rankings" in all_values


def test_comparison_ratio_formula(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    ws = wb["Model"]
    # Find a ratio formula cell
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "/" in cell.value and "$" in cell.value:
                found = True
    assert found, "No ratio formula found in Model sheet"


def test_comparison_rank_formula(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    ws = wb["Model"]
    # Find RANK formula
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "RANK" in cell.value:
                found = True
    assert found, "No RANK formula found in Model sheet"


def test_comparison_constant_values(comparison_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, comparison_spec, None, style)
    ws = wb["Model"]
    # Constant formula items should write literal 0 values
    found_zero = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 0:
                found_zero = True
    assert found_zero, "No constant value 0 found in Model sheet"


def _build_wb(spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_comparison(wb, spec, None, style)
    return wb


def _make_spec(line_items, entities):
    return ModelSpec(
        model_type="comparison",
        title="Test Comparison",
        currency="EUR",
        granularity="auto",
        start_period="2025",
        n_periods=0,
        n_history_periods=0,
        assumptions=(),
        line_items=tuple(line_items),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=tuple(entities),
        drivers=(),
    )


_ENTITIES = (
    EntityDef(key="company_a", label="Company A"),
    EntityDef(key="company_b", label="Company B"),
)


def test_comparison_subtotal_styling(style):
    spec = _make_spec(
        line_items=[
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="subtotal_rev",
                label="Subtotal Revenue",
                formula_type="sum_of_rows",
                formula_params={"addend_keys": ["revenue"]},
                is_subtotal=True,
                is_total=False,
                section="",
                format="",
            ),
        ],
        entities=_ENTITIES,
    )
    wb = _build_wb(spec, style)
    ws = wb["Model"]
    # Find the subtotal label row
    for row_idx in range(3, 20):
        if ws.cell(row=row_idx, column=1).value == "Subtotal Revenue":
            label_cell = ws.cell(row=row_idx, column=1)
            data_cell = ws.cell(row=row_idx, column=2)
            assert label_cell.fill.start_color.rgb == "00D6E4F0"
            assert data_cell.fill.start_color.rgb == "00D6E4F0"
            return
    pytest.fail("Subtotal row not found")


def test_comparison_total_styling(style):
    spec = _make_spec(
        line_items=[
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="total_rev",
                label="Total Revenue",
                formula_type="sum_of_rows",
                formula_params={"addend_keys": ["revenue"]},
                is_subtotal=False,
                is_total=True,
                section="",
                format="",
            ),
        ],
        entities=_ENTITIES,
    )
    wb = _build_wb(spec, style)
    ws = wb["Model"]
    for row_idx in range(3, 20):
        if ws.cell(row=row_idx, column=1).value == "Total Revenue":
            label_cell = ws.cell(row=row_idx, column=1)
            data_cell = ws.cell(row=row_idx, column=2)
            assert label_cell.fill.start_color.rgb == "00AED6F1"
            assert data_cell.fill.start_color.rgb == "00AED6F1"
            return
    pytest.fail("Total row not found")


def test_comparison_index_to_base_formula(style):
    spec = _make_spec(
        line_items=[
            LineItemDef(
                key="revenue",
                label="Revenue",
                formula_type="constant",
                formula_params={"value": 100},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
            LineItemDef(
                key="rev_index",
                label="Revenue Index",
                formula_type="index_to_base",
                formula_params={"value_key": "revenue", "base_entity_key": "company_a"},
                is_subtotal=False,
                is_total=False,
                section="",
                format="",
            ),
        ],
        entities=_ENTITIES,
    )
    wb = _build_wb(spec, style)
    ws = wb["Model"]
    for row_idx in range(3, 20):
        if ws.cell(row=row_idx, column=1).value == "Revenue Index":
            # Company B column (col 3) should have a formula referencing base col B
            formula = ws.cell(row=row_idx, column=3).value
            assert isinstance(formula, str)
            assert formula.startswith("=")
            assert "/" in formula
            assert "B" in formula  # base entity column letter
            return
    pytest.fail("index_to_base row not found")


def test_entity_col_index_known_key(comparison_spec):
    assert _entity_col_index(comparison_spec, "company_a") == 2
    assert _entity_col_index(comparison_spec, "company_b") == 3
    assert _entity_col_index(comparison_spec, "company_c") == 4


def test_entity_col_index_unknown_key(comparison_spec):
    assert _entity_col_index(comparison_spec, "nonexistent") is None
