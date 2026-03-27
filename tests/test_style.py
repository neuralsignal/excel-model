"""Tests for style.py."""

from dataclasses import FrozenInstanceError

import pytest
from hypothesis import given
from hypothesis import strategies as st
from openpyxl import Workbook

from excel_model.style import (
    _thin_border,
    apply_assumption_sheet_validation,
    apply_conditional_formatting,
    apply_header_style,
    get_number_format,
)

SAMPLE_STYLE_DATA = {
    "header_fill_hex": "1F3864",
    "header_font_color": "FFFFFF",
    "subtotal_fill_hex": "D6E4F0",
    "total_fill_hex": "AED6F1",
    "history_col_fill_hex": "F2F2F2",
    "section_header_fill_hex": "E8F4FD",
    "font_name": "Calibri",
    "font_size": 10,
    "number_format_currency": "#,##0",
    "number_format_percent": "0.0%",
    "number_format_integer": "#,##0",
    "number_format_number": "#,##0.00",
}


def test_style_config_creation(sample_style):
    assert sample_style.header_fill_hex == "1F3864"
    assert sample_style.font_name == "Calibri"
    assert sample_style.font_size == 10


def test_style_config_frozen(sample_style):
    with pytest.raises(FrozenInstanceError):
        sample_style.font_size = 12  # type: ignore


def test_get_number_format_currency(sample_style):
    assert get_number_format("currency", sample_style) == "#,##0"


def test_get_number_format_percent(sample_style):
    assert get_number_format("percent", sample_style) == "0.0%"


def test_get_number_format_integer(sample_style):
    assert get_number_format("integer", sample_style) == "#,##0"


def test_get_number_format_number(sample_style):
    assert get_number_format("number", sample_style) == "#,##0.00"


def test_get_number_format_invalid(sample_style):
    with pytest.raises(ValueError, match="Unknown format_type"):
        get_number_format("invalid_format", sample_style)


def test_apply_header_style(sample_style):
    wb = Workbook()
    ws = wb.active
    cell = ws["A1"]
    apply_header_style(cell, sample_style)
    # Cell should have fill applied
    assert cell.fill is not None
    assert cell.font is not None
    assert cell.font.bold is True


def test_thin_border_returns_border():
    border = _thin_border()
    assert border.bottom.style == "thin"


def test_apply_conditional_formatting_positive_is_good(sample_style):
    wb = Workbook()
    ws = wb.active
    apply_conditional_formatting(ws, "B2:C2", True, sample_style)
    cf_list = list(ws.conditional_formatting)
    assert len(cf_list) == 1
    assert len(cf_list[0].rules) == 2
    assert cf_list[0].rules[0].operator == "greaterThan"
    assert cf_list[0].rules[1].operator == "lessThan"


def test_apply_conditional_formatting_cost_item(sample_style):
    wb = Workbook()
    ws = wb.active
    apply_conditional_formatting(ws, "B2:C2", False, sample_style)
    cf_list = list(ws.conditional_formatting)
    assert len(cf_list) == 1
    assert len(cf_list[0].rules) == 2
    # For cost items, positive values get red (unfavorable)
    assert cf_list[0].rules[0].operator == "greaterThan"
    assert cf_list[0].rules[1].operator == "lessThan"


def test_apply_assumption_sheet_validation_empty():
    wb = Workbook()
    ws = wb.active
    apply_assumption_sheet_validation(ws, [], 3, 4, 3)
    assert len(ws.data_validations.dataValidation) == 0


@given(
    col=st.integers(min_value=1, max_value=26),
    row_start=st.integers(min_value=1, max_value=100),
    row_end=st.integers(min_value=1, max_value=100),
    positive_is_good=st.booleans(),
)
def test_apply_conditional_formatting_always_adds_two_rules(col, row_start, row_end, positive_is_good):
    from openpyxl.utils import get_column_letter

    from excel_model.style import StyleConfig

    lo, hi = min(row_start, row_end), max(row_start, row_end)
    letter = get_column_letter(col)
    cell_range = f"{letter}{lo}:{letter}{hi}"

    style = StyleConfig(**SAMPLE_STYLE_DATA)
    wb = Workbook()
    ws = wb.active
    apply_conditional_formatting(ws, cell_range, positive_is_good, style)
    cf_list = list(ws.conditional_formatting)
    assert len(cf_list) == 1
    assert len(cf_list[0].rules) == 2
