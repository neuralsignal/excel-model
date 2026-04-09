"""Tests for data sheet builders and validators."""

from __future__ import annotations

import pytest
from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from excel_model.config import load_style
from excel_model.data_sheet_validator import (
    validate_data_sheet_def,
    validate_sumifs_pivot_def,
)
from excel_model.exceptions import SpecValidationError
from excel_model.models.data_sheet import build_data_sheet, build_sumifs_pivot
from excel_model.spec import DataSheetDef, SumifsPivotDef


@pytest.fixture
def style():
    return load_style(None)


@pytest.fixture
def wb():
    workbook = Workbook()
    del workbook["Sheet"]
    return workbook


def _data_spec(**overrides) -> DataSheetDef:
    """Helper with sensible defaults for DataSheetDef tests."""
    defaults = dict(
        sheet_name="TEST_SHEET",
        title="Test",
        headers=("A", "B"),
        col_widths=(15.0, 15.0),
        number_formats={},
        freeze_row=2,
    )
    defaults.update(overrides)
    return DataSheetDef(**defaults)


def _pivot_spec(**overrides) -> SumifsPivotDef:
    """Helper with sensible defaults for SumifsPivotDef tests."""
    defaults = dict(
        sheet_name="SUMIFS_TEST",
        title="Test Pivot",
        row_label_headers=("Hub",),
        col_dim_values=(2023, 2024, 2025),
        data_sheet="TRANSACTIONS_LNFW",
        value_col="AO",
        row_filter_cols=("AM",),
        col_filter_col="AJ",
        append_total=True,
        append_yoy=False,
        col_widths=(28.0, 14.0, 14.0, 14.0, 16.0),
        number_format_data="#,##0",
        number_format_pct="0.0%",
        freeze_row=2,
    )
    defaults.update(overrides)
    return SumifsPivotDef(**defaults)


# ---------------------------------------------------------------------------
# DataSheetDef validation tests
# ---------------------------------------------------------------------------


def test_valid_data_sheet_def_passes():
    errors = validate_data_sheet_def(_data_spec())
    assert errors == []


def test_data_sheet_empty_headers_rejected():
    errors = validate_data_sheet_def(_data_spec(headers=(), col_widths=()))
    assert any("headers must not be empty" in e for e in errors)


def test_data_sheet_col_widths_mismatch_rejected():
    errors = validate_data_sheet_def(_data_spec(headers=("A", "B"), col_widths=(10.0,)))
    assert any("col_widths length" in e for e in errors)


def test_data_sheet_number_format_out_of_range():
    errors = validate_data_sheet_def(_data_spec(number_formats={5: "#,##0"}))
    assert any("out of range" in e for e in errors)


def test_data_sheet_negative_freeze_row():
    errors = validate_data_sheet_def(_data_spec(freeze_row=-1))
    assert any("freeze_row" in e for e in errors)


def test_data_sheet_unsafe_sheet_name():
    errors = validate_data_sheet_def(_data_spec(sheet_name="Sheet!A1"))
    assert any("unsafe characters" in e for e in errors)


# ---------------------------------------------------------------------------
# SumifsPivotDef validation tests
# ---------------------------------------------------------------------------


def test_valid_sumifs_def_passes():
    errors = validate_sumifs_pivot_def(_pivot_spec())
    assert errors == []


def test_sumifs_invalid_value_col():
    errors = validate_sumifs_pivot_def(_pivot_spec(value_col="abc"))
    assert any("value_col" in e for e in errors)


def test_sumifs_invalid_filter_col():
    errors = validate_sumifs_pivot_def(_pivot_spec(row_filter_cols=("123",)))
    assert any("row_filter_cols" in e for e in errors)


def test_sumifs_filter_cols_exceed_labels():
    errors = validate_sumifs_pivot_def(_pivot_spec(row_label_headers=("Hub",), row_filter_cols=("AM", "AN")))
    assert any("must not exceed" in e for e in errors)


def test_sumifs_unsafe_data_sheet_name():
    errors = validate_sumifs_pivot_def(_pivot_spec(data_sheet="Sheet!CMD"))
    assert any("unsafe characters" in e for e in errors)


def test_sumifs_col_widths_mismatch():
    errors = validate_sumifs_pivot_def(_pivot_spec(col_widths=(10.0,)))
    assert any("col_widths length" in e for e in errors)


def test_sumifs_empty_row_filter_cols():
    errors = validate_sumifs_pivot_def(_pivot_spec(row_filter_cols=()))
    assert any("row_filter_cols must not be empty" in e for e in errors)


# ---------------------------------------------------------------------------
# build_data_sheet tests
# ---------------------------------------------------------------------------


def test_creates_sheet_with_correct_name(wb, style):
    build_data_sheet(wb=wb, spec=_data_spec(), rows=[[1, 2], [3, 4]], style=style)
    assert "TEST_SHEET" in wb.sheetnames


def test_row_count(wb, style):
    data_rows = [[i, i * 2] for i in range(10)]
    ws = build_data_sheet(wb=wb, spec=_data_spec(), rows=data_rows, style=style)
    assert ws.max_row == 2 + len(data_rows)


def test_header_values(wb, style):
    spec = _data_spec(
        sheet_name="HDR",
        headers=("Hub", "2023", "2024", "Total"),
        col_widths=(20.0, 12.0, 12.0, 14.0),
    )
    build_data_sheet(wb=wb, spec=spec, rows=[["DE Berlin", 100, 200, 300]], style=style)
    ws = wb["HDR"]
    actual = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert actual == ["Hub", "2023", "2024", "Total"]


def test_number_format_applied(wb, style):
    spec = _data_spec(
        sheet_name="FMT",
        headers=("Name", "Amount"),
        number_formats={1: "#,##0"},
    )
    ws = build_data_sheet(wb=wb, spec=spec, rows=[["Vendor A", 12345.67]], style=style)
    assert ws.cell(row=3, column=2).number_format == "#,##0"


def test_freeze_panes_default(wb, style):
    ws = build_data_sheet(wb=wb, spec=_data_spec(sheet_name="FRZ"), rows=[[1, 2]], style=style)
    assert ws.freeze_panes == "A3"


def test_freeze_panes_custom(wb, style):
    spec = _data_spec(sheet_name="FRZ2", headers=("A",), col_widths=(10.0,), freeze_row=1)
    ws = build_data_sheet(wb=wb, spec=spec, rows=[[1]], style=style)
    assert ws.freeze_panes == "A2"


def test_col_widths(wb, style):
    widths = (25.0, 12.5, 18.0)
    spec = _data_spec(
        sheet_name="WID",
        headers=("A", "B", "C"),
        col_widths=widths,
    )
    ws = build_data_sheet(wb=wb, spec=spec, rows=[], style=style)
    for i, expected in enumerate(widths, start=1):
        assert ws.column_dimensions[get_column_letter(i)].width == expected


def test_title_in_row_1(wb, style):
    spec = _data_spec(sheet_name="TTL", headers=("X",), col_widths=(10.0,), title="My Title")
    ws = build_data_sheet(wb=wb, spec=spec, rows=[[42]], style=style)
    assert ws["A1"].value == "My Title"


def test_empty_rows(wb, style):
    spec = _data_spec(sheet_name="EMPTY")
    ws = build_data_sheet(wb=wb, spec=spec, rows=[], style=style)
    assert ws.max_row == 2


def test_returns_worksheet(wb, style):
    from openpyxl.worksheet.worksheet import Worksheet

    result = build_data_sheet(wb=wb, spec=_data_spec(sheet_name="RET"), rows=[[1, 2]], style=style)
    assert isinstance(result, Worksheet)


def test_data_sheet_rejects_invalid_spec(wb, style):
    spec = _data_spec(headers=(), col_widths=())
    with pytest.raises(SpecValidationError, match="headers must not be empty"):
        build_data_sheet(wb=wb, spec=spec, rows=[], style=style)


# ---------------------------------------------------------------------------
# Property test for build_data_sheet
# ---------------------------------------------------------------------------

_printable_text = st.text(
    alphabet=st.characters(
        whitelist_categories=("L", "N", "P", "S"),
    ),
    min_size=1,
    max_size=20,
)


@given(
    headers=st.lists(_printable_text, min_size=1, max_size=8),
    n_rows=st.integers(min_value=0, max_value=30),
)
@settings(max_examples=50)
def test_build_data_sheet_row_count_property(headers, n_rows):
    style = load_style(None)
    workbook = Workbook()
    del workbook["Sheet"]
    rows = [[None] * len(headers) for _ in range(n_rows)]
    col_widths = tuple(10.0 for _ in headers)
    spec = DataSheetDef(
        sheet_name="PROP",
        title="Property Test",
        headers=tuple(headers),
        col_widths=col_widths,
        number_formats={},
        freeze_row=2,
    )
    ws = build_data_sheet(wb=workbook, spec=spec, rows=rows, style=style)
    assert ws.max_row == 2 + n_rows


# ---------------------------------------------------------------------------
# build_sumifs_pivot tests
# ---------------------------------------------------------------------------


def test_sumifs_creates_sheet(wb, style):
    spec = _pivot_spec(sheet_name="MY_PIVOT")
    build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["DE Berlin"], ["DE Munich"], ["DE Hamburg"]], style=style)
    assert "MY_PIVOT" in wb.sheetnames


def test_sumifs_row_count(wb, style):
    row_labels = [["Hub A"], ["Hub B"], ["Hub C"]]
    ws = build_sumifs_pivot(wb=wb, spec=_pivot_spec(), row_labels=row_labels, style=style)
    assert ws.max_row == 2 + len(row_labels)


def test_sumifs_formula_in_data_cell(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"], ["DE Munich"], ["DE Hamburg"]],
        style=style,
    )
    cell_value = ws.cell(row=3, column=2).value
    assert isinstance(cell_value, str)
    assert cell_value.startswith("=SUMIFS(")
    assert "TRANSACTIONS_LNFW" in cell_value
    assert "$AO:$AO" in cell_value
    assert "$AM:$AM" in cell_value
    assert "$AJ:$AJ" in cell_value


def test_sumifs_formula_row_ref_pattern(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"], ["DE Munich"], ["DE Hamburg"]],
        style=style,
    )
    formula_row3 = ws.cell(row=3, column=2).value
    formula_row4 = ws.cell(row=4, column=2).value
    assert "$A3" in formula_row3
    assert "$A4" in formula_row4


def test_sumifs_formula_col_ref_pattern(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"], ["DE Munich"], ["DE Hamburg"]],
        style=style,
    )
    formula_col2 = ws.cell(row=3, column=2).value
    formula_col3 = ws.cell(row=3, column=3).value
    assert "B$2" in formula_col2
    assert "C$2" in formula_col3


def test_sumifs_row_labels_written(wb, style):
    row_labels = [["DE Kiel-NB"], ["DE Braunschweig"]]
    ws = build_sumifs_pivot(wb=wb, spec=_pivot_spec(), row_labels=row_labels, style=style)
    assert ws.cell(row=3, column=1).value == "DE Kiel-NB"
    assert ws.cell(row=4, column=1).value == "DE Braunschweig"


def test_sumifs_multi_label_cols(wb, style):
    spec = SumifsPivotDef(
        sheet_name="BY_CAT",
        title="By Category",
        row_label_headers=("Hub", "Account"),
        col_dim_values=(2023, 2024),
        data_sheet="TRANSACTIONS_LNFW",
        value_col="AO",
        row_filter_cols=("AM", "AN"),
        col_filter_col="AJ",
        append_total=True,
        append_yoy=False,
        col_widths=(28.0, 22.0, 14.0, 14.0, 16.0),
        number_format_data="#,##0",
        number_format_pct="0.0%",
        freeze_row=2,
    )
    ws = build_sumifs_pivot(
        wb=wb,
        spec=spec,
        row_labels=[["DE Berlin", "Maintenance"], ["DE Berlin", "Licenses"]],
        style=style,
    )
    formula = ws.cell(row=3, column=3).value
    assert "$AM:$AM" in formula
    assert "$AN:$AN" in formula
    assert "$A3" in formula
    assert "$B3" in formula


def test_sumifs_static_label_not_in_formula(wb, style):
    spec = SumifsPivotDef(
        sheet_name="VENDORS",
        title="Top Vendors",
        row_label_headers=("Vendor", "Category"),
        col_dim_values=(2023, 2024),
        data_sheet="TRANSACTIONS_LNFW",
        value_col="AO",
        row_filter_cols=("AQ",),
        col_filter_col="AJ",
        append_total=True,
        append_yoy=False,
        col_widths=(35.0, 18.0, 14.0, 14.0, 16.0),
        number_format_data="#,##0",
        number_format_pct="0.0%",
        freeze_row=2,
    )
    ws = build_sumifs_pivot(
        wb=wb,
        spec=spec,
        row_labels=[["Microsoft 365 (via reseller)", "Software/SaaS"]],
        style=style,
    )
    formula = ws.cell(row=3, column=3).value
    assert "$AQ:$AQ" in formula
    assert "$B3" not in formula
    assert ws.cell(row=3, column=2).value == "Software/SaaS"


def test_sumifs_total_column(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"], ["DE Munich"], ["DE Hamburg"]],
        style=style,
    )
    total_cell = ws.cell(row=3, column=5).value
    assert isinstance(total_cell, str)
    assert total_cell.startswith("=SUM(")
    assert "B3:D3" in total_cell


def test_sumifs_no_total(wb, style):
    spec = _pivot_spec(append_total=False, col_widths=(28.0, 14.0, 14.0, 14.0))
    ws = build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["DE Berlin"]], style=style)
    assert ws.max_column == 4


def test_sumifs_yoy_columns(wb, style):
    spec = SumifsPivotDef(
        sheet_name="YOY",
        title="YoY Test",
        row_label_headers=("Hub",),
        col_dim_values=(2023, 2024, 2025),
        data_sheet="TRANSACTIONS_LNFW",
        value_col="AO",
        row_filter_cols=("AM",),
        col_filter_col="AJ",
        append_total=True,
        append_yoy=True,
        col_widths=(28.0, 14.0, 14.0, 14.0, 16.0, 12.0, 12.0),
        number_format_data="#,##0",
        number_format_pct="0.0%",
        freeze_row=2,
    )
    ws = build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["DE Berlin"]], style=style)
    yoy_cell = ws.cell(row=3, column=6).value
    assert isinstance(yoy_cell, str)
    assert yoy_cell.startswith("=IF(")
    assert "ABS(" in yoy_cell


def test_sumifs_no_yoy(wb, style):
    spec = _pivot_spec(
        col_dim_values=(2023, 2024),
        append_total=True,
        append_yoy=False,
        col_widths=(28.0, 14.0, 14.0, 16.0),
    )
    ws = build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["DE Berlin"]], style=style)
    assert ws.max_column == 4


def test_sumifs_col_headers_in_row2(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"]],
        style=style,
    )
    assert ws.cell(row=2, column=2).value == 2023
    assert ws.cell(row=2, column=3).value == 2024
    assert ws.cell(row=2, column=4).value == 2025


def test_sumifs_freeze_panes(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"]],
        style=style,
    )
    assert ws.freeze_panes == "A3"


def test_sumifs_number_format_on_data_cells(wb, style):
    ws = build_sumifs_pivot(
        wb=wb,
        spec=_pivot_spec(),
        row_labels=[["DE Berlin"]],
        style=style,
    )
    assert ws.cell(row=3, column=2).number_format == "#,##0"


def test_sumifs_number_format_on_yoy_cells(wb, style):
    spec = SumifsPivotDef(
        sheet_name="FMTYOY",
        title="Fmt YoY",
        row_label_headers=("Hub",),
        col_dim_values=(2023, 2024),
        data_sheet="TRANSACTIONS_LNFW",
        value_col="AO",
        row_filter_cols=("AM",),
        col_filter_col="AJ",
        append_total=False,
        append_yoy=True,
        col_widths=(28.0, 14.0, 14.0, 12.0),
        number_format_data="#,##0",
        number_format_pct="0.0%",
        freeze_row=2,
    )
    ws = build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["DE Berlin"]], style=style)
    yoy_cell = ws.cell(row=3, column=4)
    assert yoy_cell.number_format == "0.0%"


def test_sumifs_rejects_invalid_spec(wb, style):
    spec = _pivot_spec(value_col="invalid")
    with pytest.raises(SpecValidationError, match="value_col"):
        build_sumifs_pivot(wb=wb, spec=spec, row_labels=[["A"]], style=style)
