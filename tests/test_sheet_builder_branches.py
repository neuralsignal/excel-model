"""Tests for _sheet_builder existing-sheet branches and build_inputs_sheet data path."""

from unittest.mock import patch

import polars as pl
from openpyxl import Workbook

from excel_model.loader import InputData
from excel_model.models._sheet_builder import (
    apply_data_cell_style,
    build_assumptions_sheet,
    build_drivers_sheet,
    build_inputs_sheet,
)
from excel_model.spec import DriverDef, InputsDef, LineItemDef, ModelSpec, ScenarioDef
from excel_model.time_engine import Period


def test_build_assumptions_sheet_reuses_existing_sheet(basic_spec, sample_style):
    """Calling build_assumptions_sheet twice reuses the existing sheet."""
    wb = Workbook()
    build_assumptions_sheet(wb, basic_spec, sample_style, scenario_prefix="")
    assert "Assumptions" in wb.sheetnames
    sheet_count_before = len(wb.sheetnames)

    rows = build_assumptions_sheet(wb, basic_spec, sample_style, scenario_prefix="")
    assert wb.sheetnames.count("Assumptions") == 1
    assert len(wb.sheetnames) == sheet_count_before
    # Still returns valid row mapping
    assert "RevenueGrowthRate" in rows


def test_build_drivers_sheet_reuses_existing_sheet(basic_spec, sample_style):
    """Calling build_drivers_sheet twice reuses the existing sheet."""
    driver = DriverDef(
        name="GrowthRate",
        label="Growth Rate",
        value=0.05,
        format="percent",
        group="Rates",
    )
    spec = ModelSpec(
        model_type=basic_spec.model_type,
        title=basic_spec.title,
        currency=basic_spec.currency,
        granularity=basic_spec.granularity,
        start_period=basic_spec.start_period,
        n_periods=basic_spec.n_periods,
        n_history_periods=basic_spec.n_history_periods,
        assumptions=basic_spec.assumptions,
        drivers=(driver,),
        line_items=basic_spec.line_items,
        metadata=basic_spec.metadata,
        scenarios=basic_spec.scenarios,
        column_groups=basic_spec.column_groups,
        inputs=basic_spec.inputs,
        entities=basic_spec.entities,
    )
    scenario = ScenarioDef(
        name="base",
        label="Base",
        assumption_overrides={},
        driver_overrides={},
    )
    wb = Workbook()
    build_drivers_sheet(wb, spec, sample_style, scenarios=(scenario,))
    assert "Drivers" in wb.sheetnames
    sheet_count_before = len(wb.sheetnames)

    rows = build_drivers_sheet(wb, spec, sample_style, scenarios=(scenario,))
    assert wb.sheetnames.count("Drivers") == 1
    assert len(wb.sheetnames) == sheet_count_before
    assert "GrowthRate" in rows


def test_build_inputs_sheet_reuses_existing_sheet(basic_spec, sample_style):
    """Calling build_inputs_sheet twice reuses the existing sheet."""
    periods = [Period(label="2023", index=0, is_history=True)]
    wb = Workbook()
    build_inputs_sheet(wb, basic_spec, None, sample_style, periods)
    assert "Inputs" in wb.sheetnames
    sheet_count_before = len(wb.sheetnames)

    build_inputs_sheet(wb, basic_spec, None, sample_style, periods)
    assert wb.sheetnames.count("Inputs") == 1
    assert len(wb.sheetnames) == sheet_count_before


def test_build_inputs_sheet_no_history_periods(basic_spec, sample_style):
    """With no history periods, sheet writes placeholder and returns empty dict."""
    wb = Workbook()
    result = build_inputs_sheet(wb, basic_spec, None, sample_style, periods=[])
    assert result == {}
    ws = wb["Inputs"]
    assert ws["A2"].value == "No history periods defined."


def test_build_inputs_sheet_writes_history_data(sample_style):
    """Provide InputData with history periods; verify cells contain expected values."""
    df = pl.DataFrame(
        {
            "period": ["2023", "2024"],
            "revenue_amount": [100.0, 200.0],
            "cogs_amount": [40.0, 80.0],
        }
    )
    inputs = InputData(
        df=df,
        period_col="period",
        value_cols=["revenue_amount", "cogs_amount"],
    )
    spec = ModelSpec(
        model_type="p_and_l",
        title="Test",
        currency="CHF",
        granularity="annual",
        start_period="2025",
        n_periods=3,
        n_history_periods=2,
        assumptions=(),
        drivers=(),
        line_items=(),
        metadata=__import__("excel_model.spec", fromlist=["MetadataDef"]).MetadataDef(
            preparer="T", date="2026-01-01", version="1.0"
        ),
        scenarios=(),
        column_groups=(),
        inputs=InputsDef(
            source="data.csv",
            period_col="period",
            sheet="",
            value_cols={"revenue": "revenue_amount", "cogs": "cogs_amount"},
        ),
        entities=(),
    )
    periods = [
        Period(label="2023", index=0, is_history=True),
        Period(label="2024", index=1, is_history=True),
    ]
    wb = Workbook()
    row_map = build_inputs_sheet(wb, spec, inputs, sample_style, periods)

    assert "revenue" in row_map
    assert "cogs" in row_map

    ws = wb["Inputs"]
    # Row 2 = headers, row 3+ = data
    # revenue row
    rev_row = row_map["revenue"]
    assert ws.cell(row=rev_row, column=1).value == "revenue"
    assert ws.cell(row=rev_row, column=2).value == 100.0  # 2023
    assert ws.cell(row=rev_row, column=3).value == 200.0  # 2024

    # cogs row
    cogs_row = row_map["cogs"]
    assert ws.cell(row=cogs_row, column=1).value == "cogs"
    assert ws.cell(row=cogs_row, column=2).value == 40.0  # 2023
    assert ws.cell(row=cogs_row, column=3).value == 80.0  # 2024


class TestApplyDataCellStyleHistoryCallCount:
    """Regression tests: apply_history_col_style must be called exactly once per history cell."""

    def _make_line_item(self, is_subtotal: bool, is_total: bool) -> LineItemDef:
        return LineItemDef(
            key="test",
            label="Test",
            formula_type="literal",
            formula_params={},
            is_subtotal=is_subtotal,
            is_total=is_total,
            section="",
            format="",
        )

    @patch("excel_model.models._sheet_builder.apply_history_col_style")
    def test_normal_history_cell_calls_history_style_once(self, mock_history, sample_style):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=100)
        li = self._make_line_item(is_subtotal=False, is_total=False)

        apply_data_cell_style(cell, li, sample_style, is_history=True)

        assert mock_history.call_count == 1

    @patch("excel_model.models._sheet_builder.apply_history_col_style")
    def test_subtotal_history_cell_does_not_call_history_style(self, mock_history, sample_style):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=100)
        li = self._make_line_item(is_subtotal=True, is_total=False)

        apply_data_cell_style(cell, li, sample_style, is_history=True)

        assert mock_history.call_count == 0

    @patch("excel_model.models._sheet_builder.apply_history_col_style")
    def test_total_history_cell_does_not_call_history_style(self, mock_history, sample_style):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=100)
        li = self._make_line_item(is_subtotal=False, is_total=True)

        apply_data_cell_style(cell, li, sample_style, is_history=True)

        assert mock_history.call_count == 0

    @patch("excel_model.models._sheet_builder.apply_history_col_style")
    def test_non_history_cell_does_not_call_history_style(self, mock_history, sample_style):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value=100)
        li = self._make_line_item(is_subtotal=False, is_total=False)

        apply_data_cell_style(cell, li, sample_style, is_history=False)

        assert mock_history.call_count == 0
