"""Tests for Budget vs Actuals model builder."""

import pytest
from openpyxl import Workbook

from excel_model.models.budget_vs_actuals import build_budget_vs_actuals
from excel_model.spec import (
    AssumptionDef,
    ColumnGroupDef,
    InputsDef,
    LineItemDef,
    MetadataDef,
    ModelSpec,
)
from excel_model.style import StyleConfig
from excel_model.time_engine import generate_periods


@pytest.fixture
def style():
    return StyleConfig(
        header_fill_hex="1F3864",
        header_font_color="FFFFFF",
        subtotal_fill_hex="D6E4F0",
        total_fill_hex="AED6F1",
        history_col_fill_hex="F2F2F2",
        section_header_fill_hex="E8F4FD",
        font_name="Calibri",
        font_size=10,
        number_format_currency="#,##0",
        number_format_percent="0.0%",
        number_format_integer="#,##0",
        number_format_number="#,##0.00",
    )


@pytest.fixture
def bva_spec():
    return ModelSpec(
        model_type="budget_vs_actuals",
        title="BvA Test",
        currency="CHF",
        granularity="monthly",
        start_period="2025-01",
        n_periods=3,
        n_history_periods=0,
        assumptions=(
            AssumptionDef(name="RevenueGrowthRate", label="Rev Growth", value=0.08, format="percent", group="Budget"),
        ),
        line_items=(
            LineItemDef(
                key="revenue_plan",
                label="Revenue (Plan)",
                formula_type="constant",
                formula_params={"value": 1000},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="revenue_actual",
                label="Revenue (Actual)",
                formula_type="constant",
                formula_params={"value": 1050},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
            LineItemDef(
                key="revenue",
                label="Revenue Variance",
                formula_type="variance",
                formula_params={"plan_key": "revenue_plan", "actual_key": "revenue_actual"},
                is_subtotal=False,
                is_total=False,
                section="Revenue",
                format="",
            ),
        ),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(
            ColumnGroupDef(key="plan", label="Plan", color_hex="D6E4F0"),
            ColumnGroupDef(key="actual", label="Actual", color_hex="FDFEFE"),
            ColumnGroupDef(key="variance", label="Variance", color_hex="FEF9E7"),
        ),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def test_bva_creates_sheets(bva_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025-01", 3, 0, "monthly")
    build_budget_vs_actuals(wb, bva_spec, None, style, periods)
    assert "Assumptions" in wb.sheetnames
    assert "Inputs" in wb.sheetnames
    assert "Model" in wb.sheetnames


def test_bva_model_has_sub_columns(bva_spec, style):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025-01", 3, 0, "monthly")
    build_budget_vs_actuals(wb, bva_spec, None, style, periods)
    ws = wb["Model"]
    # Row 3 should have sub-column labels (Plan, Actual, Variance)
    sub_labels = [ws.cell(row=3, column=c).value for c in range(2, 12)]
    assert "Plan" in sub_labels
    assert "Actual" in sub_labels
    assert "Variance" in sub_labels


def _make_bva_spec(line_items, n_periods=3, n_history_periods=0):
    """Helper to build a BVA ModelSpec with given line items."""
    return ModelSpec(
        model_type="budget_vs_actuals",
        title="BvA Test",
        currency="CHF",
        granularity="monthly",
        start_period="2025-01",
        n_periods=n_periods,
        n_history_periods=n_history_periods,
        assumptions=(
            AssumptionDef(name="RevenueGrowthRate", label="Rev Growth", value=0.08, format="percent", group="Budget"),
        ),
        line_items=tuple(line_items),
        metadata=MetadataDef(preparer="", date="", version="1.0"),
        scenarios=(),
        column_groups=(
            ColumnGroupDef(key="plan", label="Plan", color_hex="D6E4F0"),
            ColumnGroupDef(key="actual", label="Actual", color_hex="FDFEFE"),
            ColumnGroupDef(key="variance", label="Variance", color_hex="FEF9E7"),
        ),
        inputs=InputsDef(source="", period_col="period", sheet="", value_cols={}),
        entities=(),
        drivers=(),
    )


def _build_bva_wb(spec, style, n_periods=3, n_history=0):
    """Helper to build a BVA workbook and return (wb, ws)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025-01", n_periods, n_history, "monthly")
    build_budget_vs_actuals(wb, spec, None, style, periods)
    return wb, wb["Model"]


def test_bva_no_projection_periods(style):
    """n_periods=0 produces a workbook; projection col letters stay empty."""
    line_items = [
        LineItemDef(
            key="revenue_plan",
            label="Revenue (Plan)",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items, n_periods=0)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    periods = generate_periods("2025-01", 0, 0, "monthly")
    build_budget_vs_actuals(wb, spec, None, style, periods)
    assert "Model" in wb.sheetnames


def test_bva_subtotal_row_styling(style):
    """A line item with is_subtotal=True gets subtotal fill applied."""
    line_items = [
        LineItemDef(
            key="revenue_plan",
            label="Revenue (Plan)",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="revenue_subtotal",
            label="Revenue Subtotal",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=True,
            is_total=False,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items)
    _, ws = _build_bva_wb(spec, style)
    # No section header (section=""), so row 4 = first item, row 5 = subtotal
    label_cell = ws.cell(row=5, column=1)
    assert label_cell.value == "Revenue Subtotal"
    assert label_cell.fill.fgColor.rgb.endswith(style.subtotal_fill_hex)
    # Data cell in subtotal row also gets subtotal styling
    data_cell = ws.cell(row=5, column=2)
    assert data_cell.fill.fgColor.rgb.endswith(style.subtotal_fill_hex)


def test_bva_total_row_styling(style):
    """A line item with is_total=True gets total fill applied."""
    line_items = [
        LineItemDef(
            key="revenue_plan",
            label="Revenue (Plan)",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="revenue_total",
            label="Revenue Total",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=False,
            is_total=True,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items)
    _, ws = _build_bva_wb(spec, style)
    # Row 5 = total item (no section header)
    label_cell = ws.cell(row=5, column=1)
    assert label_cell.value == "Revenue Total"
    assert label_cell.fill.fgColor.rgb.endswith(style.total_fill_hex)
    # Data cell also gets total styling
    data_cell = ws.cell(row=5, column=2)
    assert data_cell.fill.fgColor.rgb.endswith(style.total_fill_hex)


def test_bva_input_ref_formula(style):
    """input_ref formula type injects line_item_key into params."""
    line_items = [
        LineItemDef(
            key="revenue_plan",
            label="Revenue (Plan)",
            formula_type="input_ref",
            formula_params={"projected_type": "constant", "value": 500},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items)
    _, ws = _build_bva_wb(spec, style)
    # With n_history=0, all periods are projection so input_ref delegates to projected_type
    cell = ws.cell(row=4, column=2)
    assert cell.value == 500


def test_bva_variance_conditional_formatting(style):
    """Variance row with positive_is_good=true gets conditional formatting rules."""
    line_items = [
        LineItemDef(
            key="revenue_plan",
            label="Revenue (Plan)",
            formula_type="constant",
            formula_params={"value": 1000},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="revenue_actual",
            label="Revenue (Actual)",
            formula_type="constant",
            formula_params={"value": 1050},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="revenue_var",
            label="Revenue Variance",
            formula_type="variance",
            formula_params={"plan_key": "revenue_plan", "actual_key": "revenue_actual", "positive_is_good": True},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items)
    _, ws = _build_bva_wb(spec, style)
    cf_rules = ws.conditional_formatting
    assert len(cf_rules) > 0
    # The variance row is row 6 (rows 4,5,6 for the 3 line items)
    rule_ranges = [str(r) for r in cf_rules]
    assert any("6" in r for r in rule_ranges)


def test_bva_variance_pct_cost_conditional_formatting(style):
    """variance_pct with positive_is_good=false (cost) gets conditional formatting."""
    line_items = [
        LineItemDef(
            key="cost_plan",
            label="Cost (Plan)",
            formula_type="constant",
            formula_params={"value": 500},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="cost_actual",
            label="Cost (Actual)",
            formula_type="constant",
            formula_params={"value": 600},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
        LineItemDef(
            key="cost_var_pct",
            label="Cost Variance %",
            formula_type="variance_pct",
            formula_params={"plan_key": "cost_plan", "actual_key": "cost_actual", "positive_is_good": False},
            is_subtotal=False,
            is_total=False,
            section="",
            format="",
        ),
    ]
    spec = _make_bva_spec(line_items)
    _, ws = _build_bva_wb(spec, style)
    cf_rules = ws.conditional_formatting
    assert len(cf_rules) > 0
    rule_ranges = [str(r) for r in cf_rules]
    assert any("6" in r for r in rule_ranges)
