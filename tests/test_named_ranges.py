"""Tests for named_ranges.py."""
from openpyxl import Workbook

from excel_model.named_ranges import get_col_letter, register_named_range


class TestGetColLetter:
    def test_single_letters(self):
        assert get_col_letter(1) == "A"
        assert get_col_letter(26) == "Z"

    def test_double_letters(self):
        assert get_col_letter(27) == "AA"
        assert get_col_letter(28) == "AB"
        assert get_col_letter(52) == "AZ"
        assert get_col_letter(53) == "BA"

    def test_column_3(self):
        assert get_col_letter(3) == "C"

    def test_column_26(self):
        assert get_col_letter(26) == "Z"


class TestRegisterNamedRange:
    def test_basic_registration(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assumptions"
        register_named_range(wb, "RevenueGrowthRate", "Assumptions", 5, 3)
        assert "RevenueGrowthRate" in wb.defined_names

    def test_reference_format(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assumptions"
        register_named_range(wb, "WACC", "Assumptions", 10, 3)
        dn = wb.defined_names["WACC"]
        assert "$C$10" in dn.attr_text

    def test_sheet_with_spaces(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "My Assumptions"
        register_named_range(wb, "TestRate", "My Assumptions", 5, 3)
        dn = wb.defined_names["TestRate"]
        assert "'My Assumptions'" in dn.attr_text

    def test_multiple_ranges(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assumptions"
        register_named_range(wb, "Rate1", "Assumptions", 3, 3)
        register_named_range(wb, "Rate2", "Assumptions", 4, 3)
        assert "Rate1" in wb.defined_names
        assert "Rate2" in wb.defined_names
